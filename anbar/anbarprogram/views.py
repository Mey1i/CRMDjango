from django.shortcuts import render,redirect
from django.contrib import messages
from . models import Brands,Clients,Chat,Departments,Expense,Orders,Planner,Permissions,Perms,Positions,Products,Staff,Suppliers,CustomUser,Settings,Messages
from django.http import HttpResponseRedirect
from django.urls import reverse
from django.core.files.storage import FileSystemStorage
from django.db.models import Q
from django.contrib.auth.models import User
from django.contrib.auth import authenticate
from django.contrib.auth import login as auth_login  
from django.contrib.auth import logout as auth_logout
from django.contrib.auth.decorators import login_required,user_passes_test
import openpyxl
from django.http import HttpResponse
import openpyxl.styles
from openpyxl.styles import Border, Side
from django.contrib.auth import update_session_auth_hash
import os
from django.db import connection
cursor = connection.cursor



# Create your views here.
def is_staff(user):
    return user.is_staff

def is_superuser(user):
    return user.is_superuser

def chat(request, id):
    mesaj = ''
    users = CustomUser.objects.exclude(id=request.user.id)
    receive_user = ''

    if id!=request.user.id:
        receive_user = CustomUser.objects.get(id=id)

    if 'ok' in request.POST:
        if request.POST['message'] != '':
            send_user = CustomUser.objects.get(id=request.user.id)
            receive_user = CustomUser.objects.get(id=id)
            message = Chat(message=request.POST['message'], send=send_user, receive=receive_user)
            message.save()
        else:
            mesaj = 'Mesaj daxil edin'
    

    user_id = request.user.id
    data = Chat.objects.filter(Q(send_id=id) & Q(receive_id=request.user.id) | Q(receive_id=id) & Q(send_id=request.user.id)).order_by('id')

    from django.db import connection, transaction
    cursor = connection.cursor()
    cursor.execute(f"SELECT * FROM anbarprogram_customuser a, anbarprogram_chat b WHERE a.id!='{request.user.id}' AND b.send_id=a.id AND b.receive_id='{request.user.id}' OR  a.id!='{request.user.id}' AND b.send_id='{request.user.id}' AND b.receive_id=a.id GROUP BY a.id")
    siyahi = cursor.fetchall()

    return render(request, 'chat.html', {'users': users, 'data': data, 'mesaj': mesaj, 'user_id': user_id, 'siyahi': siyahi, 'id': id, 'receive_user': receive_user})

# Возможно, здесь также нужен возврат HttpResponse


def mesaj(request, id):
	
	mesaj = ''
			
	if 'ok' in request.POST:
		if request.POST['message'] != '':
			g = CustomUser.objects.get(id=request.user.id)
			c = CustomUser.objects.get(id=id)
			daxilet = Chat(message=request.POST['message'], gonderen=g, qebul=c)
			daxilet.save()
		else:
			mesaj = 'Mesaj daxil edin'
	user_id = request.user.id
	data = CustomUser.objects.filter(Q(gonderen_id=id) & Q(qebul_id=request.user.id) | Q(qebul_id=id) & Q(gonderen_id=request.user.id)).order_by('id')

	from django.db import connection, transaction
	cursor = connection.cursor()
	cursor.execute(f"SELECT * FROM main_istifadeci a, main_chat b WHERE a.id!='{request.user.id}' AND b.gonderen_id=a.id AND b.qebul_id='{request.user.id}' OR  a.id!='{request.user.id}' AND b.gonderen_id='{request.user.id}' AND b.qebul_id=a.id GROUP BY a.id")
	siyahi = cursor.fetchall()

	return render(request, 'mesaj.html', {'data': data, 'mesaj': mesaj, 'user_id': user_id, 'siyahi': siyahi,'id':id})

def contact(request):
    if request.method == 'POST':
        name = request.POST['name']
        email = request.POST['email']
        phone = request.POST['phone']
        text = request.POST['text']

        if name == '' and email == '' and phone == '' and text == '':
            messages.error(request, 'You haven\'t filled all required fields!', extra_tags='alert-warning')
        else:
            message = Messages(name=name, email=email, phone=phone, message=text, accept=0)
            message.save()
            messages.success(request, 'Message has been entered into the database', extra_tags='alert-success')

    return render(request, 'contact.html')



@user_passes_test(lambda u: u.is_staff, login_url='profile')
@login_required
def look_messages(request,check=None):
    mesaj = Messages.objects.all()
    settings = Settings.objects.first()
    perinfo = Perms.objects.filter(user_id=request.user.id)
    
    user = request.user
    user_photo = user.photo
    first_name = user.first_name
    last_name = user.last_name
    email = user.email
    if request.user.is_superuser:
        vcheck = None
        echeck = None
    else:
        vcheck = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=0).count()
        echeck = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=1).count()
    admin = ''
    if request.user.is_superuser:
        admin = Permissions.objects.all()

    perinfo = Perms.objects.filter(user_id=request.user.id)
    
    data = {
        'vcheck':vcheck,
        'echeck':echeck,
        'user': user,
        'user_photo': user_photo,
        'first_name': first_name,
        'last_name': last_name,
        'email': email,
        'mesaj': mesaj,
        'settings': settings,
        'admin': admin,
        'perinfo': perinfo,
    }

    return render(request, 'messages.html', data)

@user_passes_test(lambda u: u.is_staff,login_url='profile')
@login_required
def delete_messages(request, id):
    try:
        info = Messages.objects.get(id=id)
    except Messages.DoesNotExist:
        messages.error(request, 'Message does not exist.',extra_tags='alert-warning')
        return HttpResponseRedirect(reverse('look_messages'))
    
    messages.info(request, 'Do you want to delete the message?',extra_tags='alert-success')
    mesaj = Messages.objects.all()
    settings = Settings.objects.first()
    admin = ''
    if request.user.is_superuser:
        admin = Permissions.objects.all()

    perinfo = Perms.objects.filter(user_id=request.user.id)
    user = request.user
    user_photo = user.photo
    first_name = user.first_name
    last_name = user.last_name
    email = user.email
    data = {'user':user,'user_photo':user_photo,'first_name':first_name,'last_name':last_name,'email':email,'mesaj': mesaj,'settings':settings,'admin':admin,'perinfo':perinfo,'message_id': id, 'mesaj': mesaj}
    return render(request, 'messages.html', data)


@user_passes_test(lambda u: u.is_staff,login_url='profile')
@login_required
def clean_messages(request, id):
    info = Messages.objects.get(id=id)
    info.delete()
    messages.info(request, 'Message has been deleted from the base!', extra_tags='alert-info')
    return HttpResponseRedirect(reverse('look_messages'))

@user_passes_test(lambda u: u.is_staff,login_url='profile')
@login_required
def accept_messages(request,id):
    if request.method == 'GET':
        try:
            mesaj = Messages.objects.get(id=id)
            mesaj.accept = 1
            mesaj.save()
            messages.info(request,'Message has been viewed successfully',extra_tags='alert-success')
        except Messages.DoesNotExist:
            messages.error(request,'Message does not exist!',extra_tags='alert-danger')
    else:
        messages.error(request,'Invalid request method',extra_tags='alert-warning')
    return HttpResponseRedirect(reverse('look_messages'))

@user_passes_test(lambda u: u.is_staff,login_url='profile')
@login_required
def cancel_messages(request,id):
    if request.method == 'GET':
        try:
            mesaj = Messages.objects.get(id=id)
            mesaj.accept = 0
            mesaj.save()
            messages.info(request,'Message has been canceled successfully',extra_tags='alert-success')
        except Messages.DoesNotExist:
            messages.error(request,'Message does not exist!',extra_tags='alert-danger')
    else:
        messages.error(request,'Invalid request method',extra_tags='alert-warning')
    return HttpResponseRedirect(reverse('look_messages'))


@user_passes_test(lambda u: u.is_staff, login_url='profile')
@login_required
def settings(request, check=None):
    settingss, created = Settings.objects.get_or_create()
    user = request.user
    user_photo = user.photo
    first_name = user.first_name
    last_name = user.last_name
    email = user.email
    settings = Settings.objects.first()  # Use first() to get a single instance

    user_id = request.user.id
    if request.user.is_superuser:
        vcheck = None
        echeck = None
    else:
        vcheck = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=0).count()
        echeck = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=1).count()

    admin = ''
    if request.user.is_superuser:
        admin = Permissions.objects.all()

    perinfo = Perms.objects.filter(user_id=request.user.id)
    data = {
        'vcheck': vcheck,
        'echeck': echeck,
        'admin': admin,
        'perinfo': perinfo,
        'settings': settings,
        'settingss': settingss,
        'created': created,
        'user': user,
        'first_name': first_name,
        'last_name': last_name,
        'email': email,
        'user_photo': user_photo,
    }
    return render(request, 'settings.html', data)


@user_passes_test(lambda u: u.is_superuser, login_url='profile')
@login_required
def update_settings(request, check=None):
    user = request.user
    user_photo = user.photo
    first_name = user.first_name
    last_name = user.last_name
    email = user.email
    user_id = request.user.id

    # Get or create a Settings object for the current user
    settingss, created = Settings.objects.get_or_create()
    settings = Settings.objects.first() 

    if request.user.is_superuser:
        vcheck = None
        echeck = None
    else:
        vcheck = Perms.objects.filter(item_id_id=check, user_id=user_id, item_value=0).count()
        echeck = Perms.objects.filter(item_id_id=check, user_id=user_id, item_value=1).count()

    admin = ''
    if request.user.is_superuser:
        admin = Permissions.objects.all()

    perinfo = Perms.objects.filter(user_id=request.user.id)

    section = request.POST.get('section', None)

    if request.method == 'POST':
        if 'cancel' in request.POST:
            return redirect('settings')

        if section == 'logo':
            if 'new_photo' in request.FILES:
                upload = request.FILES['new_photo']
                fs = FileSystemStorage()
                file = fs.save(upload.name, upload)
                file_url = fs.url(file)
                settings.photo = file_url

            settings.title = request.POST.get('title', '')
        elif section == 'contact':
            settings.email = request.POST.get('email', '')
            settings.phone = request.POST.get('phone', '')
            settings.address = request.POST.get('address', '')
        elif section == 'footer':
            settings.footer = request.POST.get('footer', '')

        # Save the updated settings
        settings.save()

        messages.success(request, 'Settings have been updated successfully!', extra_tags='alert-success')

        return redirect('settings')

    data = {
        'admin': admin,
        'perinfo': perinfo,
        'vcheck': vcheck,
        'echeck': echeck,
        'email': email,
        'last_name': last_name,
        'first_name': first_name,
        'settingss':settingss,
        'settings': settings,
        'user': user,
        'user_photo': user_photo,
    }
    return render(request, 'settings.html', data)


@user_passes_test(lambda u: u.is_superuser,login_url='profile')
@login_required
def manage(request,check=None):
    user_id = ''
    user = request.user
    user_photo = user.photo
    first_name = user.first_name
    last_name = user.last_name
    email = user.email
    all_users = CustomUser.objects.all()
    permissione = Permissions.objects.all()
    permsone = []
    settings = Settings.objects.first()
    if request.user.is_superuser:
        vcheck = None
        echeck = None
    else:
        vcheck = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=0).count()
        echeck = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=1).count()
    admin = ''
    if request.user.is_superuser:
        admin = Permissions.objects.all()

    perinfo = Perms.objects.filter(user_id=request.user.id)
    if 'user_id' in request.POST:
        user_id = request.POST['user_id']
        permsone = Perms.objects.filter(user_id=request.POST['user_id'])
   
    if 'ok' in request.POST:
        Perms.objects.filter(user_id=request.POST['user_id']).delete()
        perms = request.POST.getlist('secim[]')
        perms_value = request.POST.getlist('secim_value[]')
        for p in perms:
            icaze = Permissions.objects.get(id=p)
            icazeler = Perms(
                user_id=request.POST['user_id'],
                item_id=icaze,
                item_value = False
            )
            icazeler.save()

        
        for p in perms_value:
            icaze = Permissions.objects.get(id=p)
            icazeler = Perms(
                user_id=request.POST['user_id'],
                item_id=icaze,
                item_value = True
            )
            icazeler.save()
    
    return render(request, 'manage.html', {'user_photo':user_photo,'first_name':first_name,'last_name':last_name,'email':email,'admin':admin,'perinfo':perinfo,'vcheck':vcheck,'echeck':echeck,'settings':settings,'permsone':permsone,'all_users': all_users,'user_id':user_id,'permissione':permissione})

@user_passes_test(lambda u: u.is_superuser,login_url='profile')
@login_required
def permissions(request):
    if request.method == 'POST':
        x = request.POST['table']
        y = request.POST['url']
        if x !='' and y !='':
            enter = Permissions (table=x, url=y)
            enter.save()
            messages.info(request,'Information has been added into the database',extra_tags='alert-success')
        else:
            messages.info(request,'You haven\'t filled in all required information',extra_tags='alert-warning')
    info = Permissions.objects.all()
    return render(request,'permissions.html',{'info':info})

@user_passes_test(lambda u: u.is_superuser,login_url='profile')
@login_required
def perms(request):
    info = Perms.objects.all()
    return render(request, 'perms.html', {'info': info})


@user_passes_test(lambda u: u.is_staff,login_url='profile')
@login_required
def sadmin(request,check=None):
    users = CustomUser.objects.all() 
    user = request.user
    user_photo = user.photo
    first_name = user.first_name
    last_name = user.last_name
    email = user.email
    settings = Settings.objects.first()
    if request.user.is_superuser:
        vcheck = None
        echeck = None
    else:
        vcheck = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=0).count()
        echeck = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=1).count()
    admin = ''
    if request.user.is_superuser:
        admin = Permissions.objects.all()

    perinfo = Perms.objects.filter(user_id=request.user.id)
    return render(request, 'sadmin.html', {'user_photo':user_photo,'first_name':first_name,'email':email,'settings':settings,'last_name':last_name,'users': users,'vcheck':vcheck,'echeck':echeck,'admin':admin,'perinfo':perinfo,'settings':settings,})

@user_passes_test(lambda u: u.is_staff,login_url='profile')
@login_required
def delete_user(request, user_id,check=None):
    try:
        users = CustomUser.objects.all()
        settings = Settings.objects.all()
        user = request.user
        user_photo = user.photo
        first_name = user.first_name
        last_name = user.last_name
        email = user.email
        if request.user.is_superuser:
            vcheck = None
            echeck = None
        else:
            vcheck = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=0).count()
            echeck = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=1).count()
        admin = ''
        if request.user.is_superuser:
            admin = Permissions.objects.all()

        perinfo = Perms.objects.filter(user_id=request.user.id) 
        user = CustomUser.objects.get(id=user_id)
        user.delete()
        
        messages.success(request, 'User has been deleted successfully.',extra_tags='alert-success')
    except CustomUser.DoesNotExist:
        messages.error(request, 'User does not exist.',extra_tags='alert-danger')
    
    return redirect('sadmin')


@user_passes_test(lambda u: u.is_superuser,login_url='profile')
@login_required
def block_user(request, user_id,check=None):
    try:
        user = CustomUser.objects.get(id=user_id)
        settings = Settings.objects.first()
        if request.user.is_superuser:
            vcheck = None
            echeck = None
        else:
            vcheck = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=0).count()
            echeck = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=1).count()
        admin = ''
        if request.user.is_superuser:
            admin = Permissions.objects.all()

        perinfo = Perms.objects.filter(user_id=request.user.id)
        data = {'user':user,'vcheck':vcheck,'echeck':echeck,'admin':admin,'perinfo':perinfo}
        user.is_blocked = not user.is_blocked
        user.save()
        if user.is_blocked:
            messages.success(request, 'User has been blocked successfully!',extra_tags='alert-success')
        else:
            messages.success(request, 'User has been unblocked successfully!',extra_tags='alert-success')
    except CustomUser.DoesNotExist:
        messages.error(request, 'User not found.')
    return redirect('sadmin')

@user_passes_test(lambda u: u.is_superuser,login_url='profile')
@login_required
def set_admin_user(request,user_id,check=None):
    try:
        user = CustomUser.objects.get(id=user_id)
        user.is_staff = not user.is_staff
        user.save()
        if user.is_staff:
            messages.success(request,'User is admin now',extra_tags='alert-success')
        else:
            messages.success(request,'User is not admin now',extra_tags='alert-success')
    except CustomUser.DoesNotExist:
        messages.error(request,'User not found',extra_tags='alert-danger')
    return redirect('sadmin')


@user_passes_test(lambda u: u.is_staff, login_url='profile')
@login_required
def update_user(request, user_id, check=None):
    try:
        if user_id == request.user.id:  # Fix the equality check here
            user = request.user
            user_photo = user.photo
            first_name = user.first_name
            last_name = user.last_name
            email = user.email
            settings = Settings.objects.first()
        else:
            user = CustomUser.objects.get(id=user_id)
            user_photo = user.photo
            first_name = user.first_name
            last_name = user.last_name
            email = user.email
            settings = Settings.objects.first()
        if request.user.is_superuser:
            vcheck = None
            echeck = None
        else:
            vcheck = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=0).count()
            echeck = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=1).count()
        admin = ''
        if request.user.is_superuser:
            admin = Permissions.objects.all()

        perinfo = Perms.objects.filter(user_id=request.user.id)

        if request.method == 'POST':
            # Обработка формы обновления пользователя
            first_name = request.POST['first_name']
            last_name = request.POST['last_name']
            phone = request.POST['phone']
            email = request.POST['email']
            password = request.POST['password']

            # Обновление данных пользователя
            user.first_name = first_name
            user.last_name = last_name
            user.phone = phone
            user.email = email
            if password:
                user.set_password(password)
            user.save()

            # Обработка изображения профиля
            if 'new_photo' in request.FILES:
                new_photo = request.FILES['new_photo']
                fs = FileSystemStorage()
                filename = fs.save(new_photo.name, new_photo)
                user.photo = fs.url(filename)
                user.save()

            # После успешного обновления пользователя перенаправляем на страницу sadmin
            messages.success(request, 'User has been updated successfully!', extra_tags='alert-success')
            return render(request, 'sadmin.html', {'user_photo':user_photo, 'perinfo': perinfo, 'vcheck': vcheck, 'echeck': echeck,
                                                   'admin': admin,  'settings': settings, 'first_name': first_name,
                                                   'last_name': last_name,  'email': email, 'user_photo': user.photo,
                                                   'update_user': True,  'user': user})

    except CustomUser.DoesNotExist:
        messages.error(request, 'User not found.', extra_tags='alert-danger')

    # Вернуть страницу с обновленными данными пользователя
    return render(request, 'sadmin.html', {'perinfo':perinfo,'update_user': True, 'user': user, 'settings': settings,
                                            'user_photo': user.photo})


@user_passes_test(lambda u: u.is_superuser, login_url='profile')
@login_required
def menu(request, check):
    user = request.user
    user_photo = user.photo
    first_name = user.first_name
    last_name = user.last_name
    email = user.email

    # Filter Perms for the current user
    perinfo = Perms.objects.filter(user_id=request.user.id)

    settings = Settings.objects.first()

    if request.user.is_superuser:
        vcheck = None
        echeck = None
        admin = Permissions.objects.all()
    else:
        vcheck = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=0).count()
        echeck = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=1).count()
        admin = ''

    data = {'vcheck': vcheck, 'echeck': echeck, 'admin': admin, 'perinfo': perinfo, 'settings': settings,
            'user': user, 'first_name': first_name, 'last_name': last_name, 'email': email, 'user_photo': user_photo}
    return render(request, 'menu.html', data)

@login_required
def stats(request,check=None):
    #STAT START
    dsay = Departments.objects.filter(user_id = request.user.id).count()
    bsay = Brands.objects.filter(user_id=request.user.id).count()
    csay = Clients.objects.filter(user_id=request.user.id).count()
    esay = Expense.objects.filter(user_id=request.user.id).count()
    osay= Orders.objects.filter(user_id=request.user.id).count()
    tsay= Planner.objects.filter(user_id=request.user.id).count()
    psay= Positions.objects.filter(user_id=request.user.id).count()
    prosay= Products.objects.filter(user_id=request.user.id).count()
    ssay= Staff.objects.filter(user_id=request.user.id).count()
    supsay= Suppliers.objects.filter(user_id=request.user.id).count()
    user_products = Products.objects.filter(user_id=request.user.id)
    active_count= Orders.objects.filter(accept=0,user_id=request.user.id).count()
    complete_count = Orders.objects.filter(accept=1,user_id=request.user.id).count()
    active_tasks= Planner.objects.filter(accept=0,user_id=request.user.id).count()
    complete_tasks= Planner.objects.filter(accept=0,user_id=request.user.id).count()
    quantity_users = CustomUser.objects.all().count()
    user = request.user
    user_photo = user.photo
    first_name = user.first_name
    last_name = user.last_name
    email = user.email

    qazanc = 0
    for p in user_products:
        qazanc = qazanc + p.qazanc

    settings = Settings.objects.first()
    if request.user.is_superuser:
        vcheck = None
        echeck = None
    else:
        vcheck = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=0).count()
        echeck = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=1).count()
    admin = ''
    if request.user.is_superuser:
        admin = Permissions.objects.all()

    perinfo = Perms.objects.filter(user_id=request.user.id)

    # STAT END
    data = {'vcheck':vcheck,'echeck':echeck,'admin':admin,'perinfo':perinfo,'settings':settings,'bsay':bsay,'csay':csay,'dsay':dsay,'esay':esay,'osay':osay,'tsay':tsay,'psay':psay,'prosay':prosay,'ssay':ssay,'supsay':supsay,'user_products':user_products,'qazanc':qazanc,'active_count':active_count,'complete_count':complete_count,'active_tasks':active_tasks,'complete_tasks':complete_tasks,'quantity_users':quantity_users,'user':user,'first_name':first_name,'last_name':last_name,'email':email,'user_photo':user_photo}

    return render(request,'stats.html',data)

@login_required
def profile(request,check=None):
    user = request.user
    user_photo = user.photo
    first_name = user.first_name
    last_name = user.last_name
    email = user.email
    settings = Settings.objects.first()
    if request.user.is_superuser:
        vcheck = None
    else:
        vcheck = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=0).count()
    admin = ''
    if request.user.is_superuser:
        admin = Permissions.objects.all()

    perinfo = Perms.objects.filter(user_id=request.user.id)

    data = {'vcheck':vcheck,'admin':admin,'perinfo':perinfo,'settings':settings,'user':user,'first_name':first_name,'last_name':last_name,'email':email,'user_photo':user_photo}
    return render(request, 'profile.html',data)

@login_required
def update_profile(request, check=None, user_id=None):
    user = request.user
    user_photo = user.photo
    first_name = user.first_name
    last_name = user.last_name
    email = user.email
    settings = Settings.objects.first()
    user_id=request.user.id
    if request.user.is_superuser:
        vcheck = None
        echeck = None
    else:
        vcheck = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=0).count()
        echeck = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=1).count()
    admin = ''
    if request.user.is_superuser:
        admin = Permissions.objects.all()

    perinfo = Perms.objects.filter(user_id=request.user.id)
    data = {'vcheck':vcheck,'echeck':echeck,'admin':admin,'perinfo':perinfo,'settings':settings,'user':user,'first_name':first_name,'last_name':last_name,'email':email,'user_photo':user_photo}


    if request.method == 'POST':
        user.first_name = request.POST.get('first_name', '')
        user.last_name = request.POST.get('last_name', '')
        user.email = request.POST.get('email', '')
        user.phone = request.POST.get('phone', '')


    

        if 'new_photo' in request.FILES:
            upload = request.FILES['new_photo']
            fs = FileSystemStorage()
            file = fs.save(upload.name, upload)
            file_url = fs.url(file)
            user.photo = file_url
        

        password = request.POST.get('password', '')
        confirm_password = request.POST.get('confirm_password', '')



        if confirm_password:
            if not user.check_password(confirm_password):
                messages.error(request, 'Password is incorrect. Please enter your current password into the confirm password field!',extra_tags='alert-danger')
                return render(request,'profile.html',data)
            
            if password:
                user.set_password(password)
                messages.info(request, 'Passoword is correct,Profile has been updated sucessfully!',extra_tags='alert-success')
                user.save()
                return render(request,'profile.html',data)
            else:
                pass
        elif any([user.first_name, user.last_name, user.email, user.phone]):
            messages.error(request, 'Please enter your current password password into the confirm password field!',extra_tags='alert-warning')
            return render(request,'profile.html',data)
        


        user.save()
        messages.info(request, 'Profile has been updated sucessfully!',extra_tags='alert-success')
        return render(request,'profile.html',data)



    return render(request, 'profile.html', data)






@login_required
def brands(request,check=None,user_id=None):
    file_url = ''
    user = request.user
    user_photo = user.photo
    first_name = user.first_name
    last_name = user.last_name
    email = user.email
    user_id = request.user.id
    search_query = request.GET.get('search', '')  # Get the search query from the GET request
        # Filter the Brands model based on the search query
    if search_query:
        brand_filter = Q(brand__icontains=search_query)
        info = Brands.objects.filter(brand_filter).values().order_by('-id')
    else:
        info = Brands.objects.order_by('-id')

    say = Brands.objects.count()
        #STAT START
    osay = Orders.objects.count()
    bsay = Brands.objects.count()
    psay = Products.objects.count()
        #STAT END
    settings = Settings.objects.first()

    perinfo = Perms.objects.filter(user_id=request.user.id)
    if request.user.is_superuser:
        vcheck = None
        echeck = None
    else:
        vcheck = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=0).count()
        echeck = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=1).count()
    admin = ''
    if request.user.is_superuser:
        admin = Permissions.objects.all()
    # STAT END
    data = {'user_id':user_id,'admin':admin,'perinfo':perinfo,'vcheck':vcheck,'echeck':echeck,'settings':settings,'info': info, 'say': say, 'search_query': search_query,'osay':osay,'bsay':bsay,'psay':psay,'user':user,'first_name':first_name,'last_name':last_name,'email':email,'user_photo':user_photo}

    if request.method == 'POST':
        x = request.POST.get('brand', '')
        
        # Check if a file was uploaded
        if 'photo' in request.FILES:
            upload = request.FILES['photo']
            fs = FileSystemStorage()
            file = fs.save(upload.name, upload)
            file_url = fs.url(file)

        if x != '':
            check1 = Brands.objects.filter(brand=x).count()
            if check1 == 0:
                enter = Brands(brand=x, photo=file_url,user_id=request.user.id)  # Save the file URL in the model
                enter.save()
                messages.info(request, 'Brand has been added to the database!',extra_tags='alert-success')
            else:
                messages.info(request, 'Brand is already in the database',extra_tags='alert-danger')
        else:
            messages.info(request, 'You haven\'t filled in the brand field!',extra_tags='alert-warning')
    return render(request, 'brands.html', data)
    
        


@login_required
def clients(request, check=None, user_id=None):
    file_url = ''
    email_check = []
    telephone_check = []
    company_check = []
    user_id = request.user.id
    search_query = request.GET.get('search', '')  # Получаем поисковой запрос из GET-запроса

    if request.method == 'POST':
        if 'name' in request.POST and 'surname' in request.POST and 'email' in request.POST and 'telephone' in request.POST and 'company' in request.POST:
            x = request.POST['name']
            y = request.POST['surname']
            a = request.POST['email']
            z = request.POST['telephone']
            b = request.POST['company']

            if 'photo' in request.FILES:
                upload = request.FILES['photo']
                fs = FileSystemStorage()
                file = fs.save(upload.name, upload)
                file_url = fs.url(file)

            if x != '' and y != '' and a != '' and z != '' and b != '':
                email_check = Clients.objects.filter(email=a)
                telephone_check = Clients.objects.filter(telephone=z)
                company_check = Clients.objects.filter(company=b)

                if not email_check and not telephone_check and not company_check:
                    enter = Clients(name=x, surname=y, email=a, telephone=z, company=b, photo=file_url, user_id=request.user.id)
                    enter.save()
                    messages.info(request, 'Client has been added into the database!', extra_tags='alert-success')
                else:
                    if email_check:
                        messages.info(request, 'Email is already in the database!', extra_tags='alert-danger')
                    if telephone_check:
                        messages.info(request, 'Telephone is already in the database!', extra_tags='alert-danger')
                    if company_check:
                        messages.info(request, 'Company is already in the database!', extra_tags='alert-danger')
            else:
                messages.info(request, 'You haven\'t filled in all required information!', extra_tags='alert-warning')

    if search_query:
        client_filter = (
            Q(name__icontains=search_query) |
            Q(surname__icontains=search_query) |
            Q(email__icontains=search_query) |
            Q(telephone__icontains=search_query) |
            Q(company__icontains=search_query)
        )
        info = Clients.objects.filter(client_filter).values().order_by('-id')
    else:
        info = Clients.objects.order_by('-id')

    say = Clients.objects.count()
        #STAT START
    osay = Orders.objects.count()
    csay = Clients.objects.count()
    psay = Products.objects.count()


    perinfo = Perms.objects.filter(user_id=request.user.id)
    user = request.user
    user_photo = user.photo
    first_name = user.first_name
    last_name = user.last_name
    email = user.email
    settings = Settings.objects.first()
    # STAT END
    if request.user.is_superuser:
        vcheck = None
        echeck = None
    else:
        vcheck_instance = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=0).first()
        vcheck = vcheck_instance.item_id_id if vcheck_instance else None
        echeck_instance = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=1).first()
        echeck = echeck_instance.item_id_id if echeck_instance else None
        admin = ''
    if request.user.is_superuser:
        admin = Permissions.objects.all()

    perinfo = Perms.objects.filter(user_id=request.user.id)
    data = {'user_id': user_id, 'admin': admin, 'perinfo': perinfo, 'vcheck': vcheck, 'echeck': echeck, 'settings': settings, 'info': info, 'say': say, 'email_check': email_check, 'telephone_check': telephone_check, 'company_check': company_check, 'search_query': search_query, 'osay': osay, 'csay': csay, 'psay': psay, 'user': user, 'first_name': first_name, 'last_name': last_name, 'email': email, 'user_photo': user_photo, 'perinfo': perinfo}
    return render(request, 'clients.html', data)






@login_required
def departments(request,check=None, user_id=None):
    if request.method == 'POST':
        x = request.POST['department']
        if x != '':
            check1 = Departments.objects.filter(department=x).count()
            if check1 == 0:
                enter = Departments(department=x,user_id=request.user.id)
                enter.save()
                messages.info(request, 'Department has been added into database!',extra_tags='alert-success')
            else:
                messages.info(request, 'Department is already in the database!',extra_tags='alert-danger')
        else:
            messages.info(request, 'You haven\'t filled in the department field!',extra_tags='alert-warning')

    search_query = request.GET.get('search', '')  # Получаем поисковой запрос из GET-запроса

    # Фильтрация модели Departments на основе поискового запроса
    if search_query:
        department_filter = Q(department__icontains=search_query)
        info = Departments.objects.filter(department_filter).order_by('-id')
    else:
        info = Departments.objects.order_by('-id')

    say = Departments.objects.count()
        #STAT START
    dsay = Departments.objects.count()
    ssay = Staff.objects.count()
    psay = Positions.objects.count()
    
    # STAT END
    user = request.user
    first_name = user.first_name
    last_name = user.last_name
    email = user.email
    user_photo = user.photo
    user_id = request.user.id
    settings = Settings.objects.first()
    if request.user.is_superuser:
        vcheck = None
        echeck = None
    else:
        vcheck_instance = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=0).first()
        vcheck = vcheck_instance.item_id_id if vcheck_instance else None
        echeck_instance = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=1).first()
        echeck = echeck_instance.item_id_id if echeck_instance else None
        admin = ''
    if request.user.is_superuser:
        admin = Permissions.objects.all()

    perinfo = Perms.objects.filter(user_id=request.user.id)
    
    data = {'user_id':user_id,'admin':admin,'perinfo':perinfo,'vcheck':vcheck,'echeck':echeck,'settings':settings,'info': info, 'say': say, 'search_query': search_query,'dsay':dsay,'ssay':ssay,'psay':psay,'user':user,'first_name':first_name,'last_name':last_name,'email':email,'user_photo':user_photo}
    return render(request, 'departments.html', data)


@login_required
def expense(request,check=None,user_id=None):
    if request.method == 'POST':
        x = request.POST['appointment']
        y = request.POST['amount']
        if x != '' and y != '':
            enter = Expense(appointment=x, amount=y, user_id=request.user.id)
            enter.save()
            messages.info(request, 'Expense has been added ito database!',extra_tags='alert-success')
        else:
            messages.info(request, 'You haven\'t filled in all required information!',extra_tags='alert-warning')

    search_query = request.GET.get('search', '') 

    if search_query:
        expense_filter = Q(appointment__icontains=search_query) | Q(amount__icontains=search_query)
        info = Expense.objects.filter(expense_filter).order_by('-id')
    else:
        info = Expense.objects.order_by('-id')

    say = Expense.objects.count()
        #STAT START
    esay = Expense.objects.count()
    # STAT END
    user = request.user
    first_name = user.first_name
    last_name = user.last_name
    email = user.email
    user_photo = user.photo
    user_id=request.user.id
    settings = Settings.objects.first()
    if request.user.is_superuser:
        vcheck = None
        echeck = None
    else:
        vcheck_instance = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=0).first()
        vcheck = vcheck_instance.item_id_id if vcheck_instance else None
        echeck_instance = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=1).first()
        echeck = echeck_instance.item_id_id if echeck_instance else None
        admin = ''
    if request.user.is_superuser:
        admin = Permissions.objects.all()

    perinfo = Perms.objects.filter(user_id=request.user.id)

    data = {'user_id':user_id,'admin':admin,'perinfo':perinfo,'vcheck':vcheck,'echeck':echeck,'settings':settings,'info': info, 'say': say, 'search_query': search_query,'esay':esay,'user':user,'first_name':first_name,'last_name':last_name,'email':email,'user_photo':user_photo}
    return render(request, 'expense.html', data)

@login_required
def orders(request,check=None, filter_status=None,user_id=None):
    if request.method == 'POST':
        x = request.POST.get('quantity')
        if x:
            client_id = request.POST.get('client_id')
            product_id = request.POST.get('product_id')
            if client_id:
                c = Clients.objects.get(id=client_id)
                p = Products.objects.get(id=product_id)
                enter = Orders(quantity=x, client_id=c, product_id=p, accept=0, user_id=request.user.id)
                enter.save()
                messages.info(request, 'Order has been added into the database!',extra_tags='alert-success')
            else:
                messages.info(request, 'Please select a client!',extra_tags='alert-warning')
        else:
            messages.info(request, 'You did not provide quantity information!',extra_tags='alert-warning')

    search_query = request.GET.get('search', '')  # Получаем поисковой запрос из GET-запроса

    if filter_status == "active":
        order_filter = Q(accept=0)
    elif filter_status == "complete":
        order_filter = Q(accept=1)
    else:
        order_filter = Q()

    if search_query:
        orders = Orders.objects.filter(order_filter)
        order_filter = (
            Q(quantity__icontains=search_query) |
            Q(client_id__name__icontains=search_query) |
            Q(client_id__surname__icontains=search_query) |
            Q(product_id__brand_id__brand__icontains=search_query) |
            Q(product_id__name__icontains=search_query) |
            Q(product_id__quantity__icontains=search_query)
        )
        info = orders.filter(order_filter).order_by('-id')
    else:
        info = Orders.objects.filter(order_filter).order_by('-id')

    say = Orders.objects.count()
    cinfo = Clients.objects.all()
    pinfo = Products.objects.all()
    active_count = Orders.objects.filter(accept=0).count()
    completed_count = Orders.objects.filter(accept=1).count()
    #STAT START
    osay = Orders.objects.count()
    bsay = Brands.objects.count()
    csay = Clients.objects.count()
    psay = Products.objects.count()
    # STAT END
    user = request.user
    user_id=request.user.id
    first_name = user.first_name
    last_name = user.last_name
    email = user.email
    user_photo = user.photo
    settings = Settings.objects.first()
    if request.user.is_superuser:
        vcheck = None
        echeck = None
    else:
        vcheck_instance = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=0).first()
        vcheck = vcheck_instance.item_id_id if vcheck_instance else None
        echeck_instance = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=1).first()
        echeck = echeck_instance.item_id_id if echeck_instance else None
        admin = ''
    if request.user.is_superuser:
        admin = Permissions.objects.all()

    perinfo = Perms.objects.filter(user_id=request.user.id)
    
    data = {'user_id':user_id,'admin':admin,'perinfo':perinfo,'vcheck':vcheck,'echeck':echeck,'settings':settings,'info': info, 'say': say, 'cinfo': cinfo, 'pinfo': pinfo, 'active_count': active_count, 'completed_count': completed_count, 'search_query': search_query,'osay':osay,'bsay':bsay,'csay':csay,'psay':psay,'user':user,'first_name':first_name,'last_name':last_name,'email':email,'user_photo':user_photo}
    return render(request, 'orders.html', data)

@login_required
def planner(request,check=None, filter_status=None,user_id=None):
    if request.method == 'POST':
        x = request.POST['task']
        y = request.POST['dtask']
        z = request.POST['ttask']
        staff_id = request.POST['staff_id']

        if x and y and z and staff_id:
            try:
                s = Staff.objects.get(id=staff_id)

                enter = Planner(task=x, dtask=y, ttask=z, staff_id=s, accept=0, user_id=request.user.id)
                enter.save()

                messages.info(request, 'Task has been added to the database!',extra_tags='alert-success')
            except Staff.DoesNotExist:
                messages.error(request, 'Staff not found. Please select a valid staff.',extra_tags='alert-danger')
        else:
            messages.info(request, 'You didn\'t fill in all the required information.',extra_tags='alert-warning')

    search_query = request.GET.get('search', '')  # Получаем поисковой запрос из GET-запроса

    if filter_status == "active":
        planner_filter = Q(accept=0)
    elif filter_status == "complete":
        planner_filter = Q(accept=1)
    else:
        planner_filter = Q()

    if search_query:
        planner = Planner.objects.filter(planner_filter)
        planner_filter = (
            Q(task__icontains=search_query) |
            Q(dtask__icontains=search_query) |
            Q(ttask__icontains=search_query) |
            Q(staff_id__name__icontains=search_query) |
            Q(staff_id__surname__icontains=search_query)
        )
        info = planner.filter(planner_filter).order_by('-id')
    else:
        info = Planner.objects.filter(planner_filter).order_by('-id')

    say = Planner.objects.count()
    sinfo = Staff.objects.all()
    active_count = Planner.objects.filter(accept=0).count()
    completed_count = Planner.objects.filter(accept=1).count()
    #STAT START
    dsay = Departments.objects.count()
    psay = Positions.objects.count()
    ssay = Staff.objects.count()
    # STAT END
    user = request.user
    first_name = user.first_name
    last_name = user.last_name
    email = user.email
    user_photo = user.photo
    settings = Settings.objects.first()
    user_id=request.user.id
    if request.user.is_superuser:
        vcheck = None
        echeck = None
    else:
        vcheck_instance = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=0).first()
        vcheck = vcheck_instance.item_id_id if vcheck_instance else None
        echeck_instance = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=1).first()
        echeck = echeck_instance.item_id_id if echeck_instance else None
        admin = ''
    if request.user.is_superuser:
        admin = Permissions.objects.all()

    perinfo = Perms.objects.filter(user_id=request.user.id)

    data = {'user_id':user_id,'admin':admin,'perinfo':perinfo,'vcheck':vcheck,'echeck':echeck,'settings':settings,'user_photo':user_photo,'settings':settings,'info': info, 'say': say, 'sinfo': sinfo, 'active_count': active_count, 'completed_count': completed_count, 'search_query': search_query,'dsay':dsay,'psay':psay,'ssay':ssay,'user':user,'first_name':first_name,'last_name':last_name,'email':email}
    return render(request, 'planner.html', data)

@login_required
def positions(request,check=None,user_id=None):
    settings = Settings.objects.first()
    if request.user.is_superuser:
        vcheck = None
        echeck = None
    else:
        vcheck_instance = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=0).first()
        vcheck = vcheck_instance.item_id_id if vcheck_instance else None
        echeck_instance = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=1).first()
        echeck = echeck_instance.item_id_id if echeck_instance else None
        admin = ''
    if request.user.is_superuser:
        admin = Permissions.objects.all()

    perinfo = Perms.objects.filter(user_id=request.user.id)
    if request.method == 'POST':
        x = request.POST['positions']
        department_id = request.POST['department_id']
        if x:
            try:
                d = Departments.objects.get(id=department_id)
                check = Positions.objects.filter(positions=x, department_id=d).count()
                if check == 0:
                    enter = Positions(positions=x, department_id=d, user_id=request.user.id)
                    enter.save()
                    messages.info(request, 'Position has been added to the database!',extra_tags='alert-success')
                else:
                    messages.info(request, 'Position is already in the database!',extra_tags='alert-danger')
            except Departments.DoesNotExist:
                messages.info(request, 'Department not found. Please select a valid department.',extra_tags='alert-danger')
        else:
            messages.info(request, 'You didn\'t fill in all the required information.',extra_tags='alert-warning')

    search_query = request.GET.get('search', '')  # Получаем поисковой запрос из GET-запроса

    if search_query:
        position_filter = Q(positions__icontains=search_query) | Q(department_id__department__icontains=search_query)
        info = Positions.objects.filter(position_filter).order_by('-id')
    else:
        info = Positions.objects.order_by('-id')

    say = Positions.objects.count()
    dinfo = Departments.objects.all()
        #STAT START
    dsay = Departments.objects.count()
    psay = Positions.objects.count()
    ssay = Staff.objects.count()
    # STAT END
    user = request.user
    user_id=request.user.id
    first_name = user.first_name
    last_name = user.last_name
    email = user.email
    user_photo = user.photo


    data = {'user_id':user_id,'admin':admin,'perinfo':perinfo,'vcheck':vcheck,'echeck':echeck,'settings':settings,'settings':settings,'info': info, 'say': say, 'dinfo': dinfo, 'search_query': search_query,'dsay':dsay,'ssay':ssay,'psay':psay,'user':user,'first_name':first_name,'last_name':last_name,'email':email,'user_photo':user_photo}
    return render(request, 'positions.html', data)

@login_required
def products(request,check=None, user_id=None):
    if request.method == 'POST':
        x = request.POST['name']
        y = request.POST['purchase']
        z = request.POST['sale']
        a = request.POST['quantity']
        brand_id = request.POST['brand_id']
        supplier_id = request.POST['supplier_id']

        if 'photo' in request.FILES:
            upload = request.FILES['photo']
            fs = FileSystemStorage()
            file = fs.save(upload.name, upload)
            file_url = fs.url(file)
        else:
            file_url = ''

        if x and y and z and a and brand_id and supplier_id:
            try:
                b = Brands.objects.get(id=brand_id)
                s = Suppliers.objects.get(id=supplier_id)
                enter = Products(brand_id=b, supplier_id=s, name=x, purchase=y, sale=z, quantity=a, photo=file_url, user_id=request.user.id)
                enter.save()
                messages.info(request, 'Product has been added into the database!',extra_tags='alert-success')
            except Brands.DoesNotExist:
                messages.error(request, 'Brand not found. Please select a valid brand.',extra_tags='alert-danger')
            except Suppliers.DoesNotExist:
                messages.error(request, 'Supplier not found. Please select a valid supplier.',extra_tags='alert-danger')
        else:
            messages.error(request, 'You haven\'t filled in all required information!',extra_tags='alert-warning')
    
    search_query = request.GET.get('search', '')  # Получаем поисковой запрос из GET-запроса


    if search_query:
        product_filter &= (
            Q(name__icontains=search_query) |
            Q(purchase__icontains=search_query) |
            Q(sale__icontains=search_query) |
            Q(quantity__icontains=search_query) |
            Q(brand_id__brand__icontains=search_query) |
            Q(supplier_id__firm__icontains=search_query)
        )
        info = Products.objects.filter(product_filter).order_by('-id')
    else:
        info = Products.objects.order_by('-id')
        
    say = info.count()
    binfo = Brands.objects.all()
    sinfo = Suppliers.objects.all()
    #STAT START
    osay = Orders.objects.count()
    bsay = Brands.objects.count()
    csay = Clients.objects.count()
    psay = Products.objects.count()
    # STAT END
    user = request.user
    user_id=request.user.id
    first_name = user.first_name
    last_name = user.last_name
    email = user.email
    user_photo = user.photo
    settings = Settings.objects.first()
    if request.user.is_superuser:
        vcheck = None
        echeck = None
    else:
        vcheck_instance = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=0).first()
        vcheck = vcheck_instance.item_id_id if vcheck_instance else None
        echeck_instance = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=1).first()
        echeck = echeck_instance.item_id_id if echeck_instance else None
        admin = ''
    if request.user.is_superuser:
        admin = Permissions.objects.all()

    perinfo = Perms.objects.filter(user_id=request.user.id)

    data = {'user_id':user_id,'admin':admin,'perinfo':perinfo,'vcheck':vcheck,'echeck':echeck,'settings':settings,'settings':settings,'info': info, 'say': say, 'binfo': binfo, 'sinfo': sinfo, 'search_query': search_query,'osay':osay,'bsay':bsay,'csay':csay,'psay':psay,'user':user,'first_name':first_name,'last_name':last_name,'email':email,'user_photo':user_photo}
    return render(request, 'products.html', data)


@login_required
def staff(request,check=None,user_id=None):

    #STAT START
    dsay = Departments.objects.count()
    ssay = Staff.objects.count()
    psay = Products.objects.count()
    # STAT END
    user = request.user
    user_id = request.user.id
    first_name = user.first_name
    last_name = user.last_name
    email = user.email
    user_photo = user.photo
    settings = Settings.objects.first()
    if request.user.is_superuser:
        vcheck = None
        echeck = None
        admin = Permissions.objects.all()
    else:
        vcheck_instance = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=0).first()
        vcheck = vcheck_instance.item_id_id if vcheck_instance else None
        echeck_instance = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=1).first()
        echeck = echeck_instance.item_id_id if echeck_instance else None
        admin = ''


        

    perinfo = Perms.objects.filter(user_id=request.user.id)
    if request.method == 'POST':
        x = request.POST['name']
        y = request.POST['surname']
        z = request.POST['email']
        a = request.POST['telephone']
        b = request.POST['salary']
        c = request.POST['work']
        position_id = request.POST['position_id']
        
        if 'photo' in request.FILES:
            upload = request.FILES['photo']
            fs = FileSystemStorage()
            file = fs.save(upload.name, upload)
            file_url = fs.url(file)
        else:
            file_url = ''

    
# STAT END


        if x and y and z and a and b and c:
            p = Positions.objects.get(id=position_id)
            if Staff.objects.filter(email=z, telephone=a).count() == 0:
                enter = Staff(position_id=p, name=x, surname=y, email=z, telephone=a, salary=b, work=c, photo=file_url, user_id=request.user.id)
                enter.save()
                messages.info(request, 'Staff has been added to the database!',extra_tags='alert-success')
            else:
                messages.info(request, 'Email or Telephone is already in the database!',extra_tags='alert-danger')
        else:
            messages.info(request, 'You did not provide all the required information!',extra_tags='alert-warning')


    search_query = request.GET.get('search', '')  # Получаем поисковой запрос из GET-запроса

    if search_query:
        staff_filter = (
            Q(name__icontains=search_query) |
            Q(surname__icontains=search_query) |
            Q(email__icontains=search_query) |
            Q(telephone__icontains=search_query) |
            Q(salary__icontains=search_query) |
            Q(work__icontains=search_query) |
            Q(position_id__positions__icontains=search_query)
        )
        info = Staff.objects.filter(staff_filter).order_by('-id')
    else:
        info = Staff.objects.order_by('-id')
    pinfo = Positions.objects.all()
    say = info.count()





    data = {'user_id':user_id,'admin':admin,'perinfo':perinfo,'vcheck':vcheck,'echeck':echeck,'settings':settings,'settings':settings,'info': info, 'say': say, 'pinfo': pinfo, 'search_query': search_query,'dsay':dsay,'ssay':ssay,'psay':psay,'user':user,'first_name':first_name,'last_name':last_name,'email':email,'user_photo':user_photo}
    return render(request, 'staff.html', data)


@login_required
def suppliers(request,check=None,user_id=None):
    firm_check = []
    email_check = []
    telephone_check = []
    address_check = []
    settings = Settings.objects.first()
    if request.user.is_superuser:
        vcheck = None
        echeck = None
    else:
        vcheck_instance = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=0).first()
        vcheck = vcheck_instance.item_id_id if vcheck_instance else None
        echeck_instance = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=1).first()
        echeck = echeck_instance.item_id_id if echeck_instance else None
        admin = ''
    if request.user.is_superuser:
        admin = Permissions.objects.all()

    perinfo = Perms.objects.filter(user_id=request.user.id)

    if request.method == 'POST':
        x = request.POST['firm']
        y = request.POST['name']
        z = request.POST['surname']
        a = request.POST['email']
        b = request.POST['telephone']
        c = request.POST['address']

        if 'photo' in request.FILES:
            upload = request.FILES['photo']
            fs = FileSystemStorage()
            file = fs.save(upload.name, upload)
            file_url = fs.url(file)
        else:
            file_url = ''

        if x and y and z and a and b and c:
            firm_check = Suppliers.objects.filter(firm=x)
            email_check = Suppliers.objects.filter(email=a)
            telephone_check = Suppliers.objects.filter(telephone=b)
            address_check = Suppliers.objects.filter(address=c)
            if not firm_check and not email_check and not telephone_check and not address_check:
                enter = Suppliers(firm=x, name=y, surname=z, email=a, telephone=b, address=c, photo=file_url, user_id=request.user.id)
                enter.save()
                messages.info(request, 'Supplier has been added to the database!',extra_tags='alert-success')
            else:
                if firm_check:
                    messages.info(request, 'Firm is already entered into the database!',extra_tags='alert-danger')
                if email_check:
                    messages.info(request, 'Email is already entered into the database!',extra_tags='alert-danger')
                if telephone_check:
                    messages.info(request, 'Telephone is already entered into the database!',extra_tags='alert-danger')
                if address_check:
                    messages.info(request, 'Address is already entered into the database!',extra_tags='alert-danger')
        else:
            messages.info(request, 'You haven\'t filled in all required information!',extra_tags='alert-warning')

    search_query = request.GET.get('search', '')  # Получаем поисковой запрос из GET-запроса

    if search_query:
        supplier_filter = (
            Q(firm__icontains=search_query) |
            Q(name__icontains=search_query) |
            Q(surname__icontains=search_query) |
            Q(email__icontains=search_query) |
            Q(telephone__icontains=search_query) |
            Q(address__icontains=search_query)
        )
        info = Suppliers.objects.filter(supplier_filter).order_by('-id')
    else:
        info = Suppliers.objects.order_by('-id')

    say = Suppliers.objects.count()
        #STAT START
    osay = Orders.objects.count()
    bsay = Brands.objects.count()
    csay = Clients.objects.count()
    psay = Products.objects.count()
    # STAT END
    user = request.user
    user_id = request.user.id
    first_name = user.first_name
    last_name = user.last_name
    email = user.email
    user_photo = user.photo


    data = {'user_id':user_id,'admin':admin,'perinfo':perinfo,'vcheck':vcheck,'echeck':echeck,'settings':settings,'settings':settings,'info': info, 'say': say, 'search_query': search_query,'osay':osay,'bsay':bsay,'csay':csay,'psay':psay,'user':user,'first_name':first_name,'last_name':last_name,'email':email,'user_photo':user_photo}
    return render(request, 'suppliers.html', data)


#DELETE FUNCTIONS START
@login_required
def delete(request, id, check=None, user_id=None):
    echeck = None
    vcheck = None
    user_id = request.user.id
    info = None

    if not request.user.is_superuser:
        echeck = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=1).count()
        vcheck = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=0).count()

    admin = ''
    if request.user.is_superuser:
        admin = Permissions.objects.all()

    try:
        info = Brands.objects.get(id=id)
        info.delete()
        messages.success(request, 'Brand has been deleted from the database successfully!',extra_tags='alert-success')

        if request.user.is_superuser:
            return HttpResponseRedirect(reverse('brands'))  # Redirect to 'brands' for superuser
        else:
            return HttpResponseRedirect(reverse('brands', args=[check, id]))  # Redirect to 'brands/check/id' for non-superuser

    except Brands.DoesNotExist:
        messages.error(request, 'Brand has been not detected!',extra_tags='alert-warning')

    data = {'echeck': echeck, 'vcheck': vcheck, 'user_id': user_id, 'admin': admin, 'info': info}
    return HttpResponseRedirect(reverse('brands'), {'data': data})



@login_required
def delete1(request, id, check=None, user_id=None):
    user_id = request.user.id
    if request.user.is_superuser:
        vcheck = None
        echeck = None
    else:
        vcheck_instance = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=0).first()
        vcheck = vcheck_instance.item_id_id if vcheck_instance else None
        echeck_instance = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=1).first()
        echeck = echeck_instance.item_id_id if echeck_instance else None
        admin = ''
    if request.user.is_superuser:
        admin = Permissions.objects.all()


    try:
        info = Clients.objects.get(id=id)
        info.delete()
        messages.success(request, 'Client has been deleted from the database successfully!',extra_tags='alert-success')

        if request.user.is_superuser:
            return HttpResponseRedirect(reverse('clients'))  # Redirect to 'brands' for superuser
        else:
            return HttpResponseRedirect(reverse('clients', args=[check, id]))  # Redirect to 'brands/check/id' for non-superuser

    except Clients.DoesNotExist:
        messages.error(request, 'Client has been not detected!',extra_tags='alert-warning')

    data = {'echeck': echeck, 'vcheck': vcheck, 'user_id': user_id, 'admin': admin, 'info': info}
    return HttpResponseRedirect(reverse('clients'), {'data': data})

@login_required
def delete2(request,id,check=None , user_id=None):
    user_id = request.user.id
    if request.user.is_superuser:
        vcheck = None
        echeck = None
    else:
        vcheck_instance = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=0).first()
        vcheck = vcheck_instance.item_id_id if vcheck_instance else None
        echeck_instance = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=1).first()
        echeck = echeck_instance.item_id_id if echeck_instance else None
        admin = ''
    if request.user.is_superuser:
        admin = Permissions.objects.all()
    try:
        info = Departments.objects.get(id=id)
        info.delete()
        messages.success(request, 'Department has been deleted from the database successfully!',extra_tags='alert-success')

        if request.user.is_superuser:
            return HttpResponseRedirect(reverse('departments'))  # Redirect to 'brands' for superuser
        else:
            return HttpResponseRedirect(reverse('departments', args=[check, id]))  # Redirect to 'brands/check/id' for non-superuser

    except Departments.DoesNotExist:
        messages.error(request, 'Department has been not detected!',extra_tags='alert-warning')

    data = {'echeck': echeck, 'vcheck': vcheck, 'user_id': user_id, 'admin': admin, 'info': info}
    return HttpResponseRedirect(reverse('departments'), {'data': data})

@login_required
def delete3(request,id,check=None , user_id=None):
    user_id = request.user.id
    if request.user.is_superuser:
        vcheck = None
        echeck = None
    else:
        vcheck_instance = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=0).first()
        vcheck = vcheck_instance.item_id_id if vcheck_instance else None
        echeck_instance = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=1).first()
        echeck = echeck_instance.item_id_id if echeck_instance else None
        admin = ''
    if request.user.is_superuser:
        admin = Permissions.objects.all()
    try:
        info = Expense.objects.get(id=id)
        info.delete()
        messages.success(request, 'Expense has been deleted from the database successfully!',extra_tags='alert-success')

        if request.user.is_superuser:
            return HttpResponseRedirect(reverse('expense'))  # Redirect to 'brands' for superuser
        else:
            return HttpResponseRedirect(reverse('expense', args=[check, id]))  # Redirect to 'brands/check/id' for non-superuser

    except Expense.DoesNotExist:
        messages.error(request, 'Expense has been not detected!',extra_tags='alert-warning')

    data = {'echeck': echeck, 'vcheck': vcheck, 'user_id': user_id, 'admin': admin, 'info': info}
    return HttpResponseRedirect(reverse('expense'), {'data': data})

@login_required
def delete4(request,id,check=None , user_id=None):
    user_id = request.user.id
    if request.user.is_superuser:
        vcheck = None
        echeck = None
    else:
        vcheck_instance = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=0).first()
        vcheck = vcheck_instance.item_id_id if vcheck_instance else None
        echeck_instance = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=1).first()
        echeck = echeck_instance.item_id_id if echeck_instance else None
        admin = ''
    if request.user.is_superuser:
        admin = Permissions.objects.all()
    try:
        info = Orders.objects.get(id=id)
        info.delete()
        messages.success(request, 'Orders has been deleted from the database successfully!',extra_tags='alert-success')

        if request.user.is_superuser:
            return HttpResponseRedirect(reverse('orders'))  # Redirect to 'brands' for superuser
        else:
            return HttpResponseRedirect(reverse('orders', args=[check, id]))  # Redirect to 'brands/check/id' for non-superuser

    except Orders.DoesNotExist:
        messages.error(request, 'Orders has been not detected!',extra_tags='alert-warning')

    data = {'echeck': echeck, 'vcheck': vcheck, 'user_id': user_id, 'admin': admin, 'info': info}
    return HttpResponseRedirect(reverse('orders'), {'data': data})

@login_required
def delete5(request,id,check=None , user_id=None):
    user_id = request.user.id
    if request.user.is_superuser:
        vcheck = None
        echeck = None
    else:
        vcheck_instance = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=0).first()
        vcheck = vcheck_instance.item_id_id if vcheck_instance else None
        echeck_instance = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=1).first()
        echeck = echeck_instance.item_id_id if echeck_instance else None
        admin = ''
    if request.user.is_superuser:
        admin = Permissions.objects.all()
    try:
        info = Planner.objects.get(id=id)
        info.delete()
        messages.success(request, 'Task has been deleted from the database successfully!',extra_tags='alert-success')

        if request.user.is_superuser:
            return HttpResponseRedirect(reverse('planner'))  # Redirect to 'brands' for superuser
        else:
            return HttpResponseRedirect(reverse('planner', args=[check, id]))  # Redirect to 'brands/check/id' for non-superuser

    except Planner.DoesNotExist:
        messages.error(request, 'Task has been not detected!',extra_tags='alert-warning')

    data = {'echeck': echeck, 'vcheck': vcheck, 'user_id': user_id, 'admin': admin, 'info': info}
    return HttpResponseRedirect(reverse('planner'), {'data': data})

@login_required
def delete6(request,id,check=None , user_id=None):
    user_id = request.user.id
    if request.user.is_superuser:
        vcheck = None
        echeck = None
    else:
        vcheck_instance = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=0).first()
        vcheck = vcheck_instance.item_id_id if vcheck_instance else None
        echeck_instance = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=1).first()
        echeck = echeck_instance.item_id_id if echeck_instance else None
        admin = ''
    if request.user.is_superuser:
        admin = Permissions.objects.all()
    try:
        info = Positions.objects.get(id=id)
        info.delete()
        messages.success(request, 'Position has been deleted from the database successfully!',extra_tags='alert-success')

        if request.user.is_superuser:
            return HttpResponseRedirect(reverse('positions'))  # Redirect to 'brands' for superuser
        else:
            return HttpResponseRedirect(reverse('positions', args=[check, id]))  # Redirect to 'brands/check/id' for non-superuser

    except Positions.DoesNotExist:
        messages.error(request, 'Position has been not detected!',extra_tags='alert-warning')

    data = {'echeck': echeck, 'vcheck': vcheck, 'user_id': user_id, 'admin': admin, 'info': info}
    return HttpResponseRedirect(reverse('expense'), {'data': data})

@login_required
def delete7(request,id,check=None , user_id=None):
    user_id = request.user.id
    if request.user.is_superuser:
        vcheck = None
        echeck = None
    else:
        vcheck_instance = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=0).first()
        vcheck = vcheck_instance.item_id_id if vcheck_instance else None
        echeck_instance = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=1).first()
        echeck = echeck_instance.item_id_id if echeck_instance else None
        admin = ''
    if request.user.is_superuser:
        admin = Permissions.objects.all()
    try:
        info = Products.objects.get(id=id)
        info.delete()
        messages.success(request, 'Product has been deleted from the database successfully!',extra_tags='alert-success')

        if request.user.is_superuser:
            return HttpResponseRedirect(reverse('products'))  # Redirect to 'brands' for superuser
        else:
            return HttpResponseRedirect(reverse('products', args=[check, id]))  # Redirect to 'brands/check/id' for non-superuser

    except Products.DoesNotExist:
        messages.error(request, 'Product has been not detected!',extra_tags='alert-warning')

    data = {'echeck': echeck, 'vcheck': vcheck, 'user_id': user_id, 'admin': admin, 'info': info}
    return HttpResponseRedirect(reverse('products'), {'data': data})

@login_required
def delete8(request,id,check=None , user_id=None):
    user_id = request.user.id
    if request.user.is_superuser:
        vcheck = None
        echeck = None
    else:
        vcheck_instance = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=0).first()
        vcheck = vcheck_instance.item_id_id if vcheck_instance else None
        echeck_instance = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=1).first()
        echeck = echeck_instance.item_id_id if echeck_instance else None
        admin = ''
    if request.user.is_superuser:
        admin = Permissions.objects.all()
    try:
        info = Staff.objects.get(id=id)
        info.delete()
        messages.success(request, 'Staff has been deleted from the database successfully!',extra_tags='alert-success')

        if request.user.is_superuser:
            return HttpResponseRedirect(reverse('staff'))  # Redirect to 'brands' for superuser
        else:
            return HttpResponseRedirect(reverse('staff', args=[check, id]))  # Redirect to 'brands/check/id' for non-superuser

    except Staff.DoesNotExist:
        messages.error(request, 'Staff has been not detected!',extra_tags='alert-warning')

    data = {'echeck': echeck, 'vcheck': vcheck, 'user_id': user_id, 'admin': admin, 'info': info}
    return HttpResponseRedirect(reverse('staff'), {'data': data})

@login_required
def delete9(request,id,check=None , user_id=None):
    user_id = request.user.id
    if request.user.is_superuser:
        vcheck = None
        echeck = None
    else:
        vcheck_instance = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=0).first()
        vcheck = vcheck_instance.item_id_id if vcheck_instance else None
        echeck_instance = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=1).first()
        echeck = echeck_instance.item_id_id if echeck_instance else None
        admin = ''
    if request.user.is_superuser:
        admin = Permissions.objects.all()
    try:
        info = Suppliers.objects.get(id=id)
        info.delete()
        messages.success(request, 'Supplier has been deleted from the database successfully!',extra_tags='alert-success')

        if request.user.is_superuser:
            return HttpResponseRedirect(reverse('suppliers'))  # Redirect to 'brands' for superuser
        else:
            return HttpResponseRedirect(reverse('suppliers', args=[check, id]))  # Redirect to 'brands/check/id' for non-superuser

    except Suppliers.DoesNotExist:
        messages.error(request, 'Supplier has been not detected!',extra_tags='alert-warning')

    data = {'echeck': echeck, 'vcheck': vcheck, 'user_id': user_id, 'admin': admin, 'info': info}
    return HttpResponseRedirect(reverse('suppliers'), {'data': data})
#DELETE FUNCTIONS END

#UDALIT FUNCTIONS START
@login_required
def udalit(request,id,check=None,user_id=None):
    messages.info(request, 'Do you want to delete the brand?',extra_tags='alert-info')
    info = Brands.objects.order_by('-id')
    say = Brands.objects.count()
         #STAT START
    osay = Orders.objects.count()
    bsay = Brands.objects.count()
    csay = Clients.objects.count()
    psay = Products.objects.count()
    # STAT END
    user = request.user
    user_id = request.user.id
    user_photo = user.photo
    first_name = user.first_name
    last_name = user.last_name
    email = user.email
    settings = Settings.objects.first()
    perinfo = Perms.objects.filter(user_id=request.user.id)
    if request.user.is_superuser:
        vcheck = None
        echeck = None
    else:
        vcheck = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=0).count()
        echeck = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=1).count()
    admin = ''
    if request.user.is_superuser:
        admin = Permissions.objects.all()
    data = {'user_id':user_id,'admin':admin,'perinfo':perinfo,'vcheck':vcheck,'echeck':echeck,'settings':settings,'info':info,'say':say,'delete_id':id,'osay':osay,'bsay':bsay,'csay':csay,'psay':psay,'user':user,'first_name':first_name,'last_name':last_name,'email':email,'user_photo':user_photo}
    return render(request,'brands.html',data)

@login_required
def udalit1(request,id,check=None):
    messages.info(request, 'Do you want to delete the client?',extra_tags='alert-info')
    info = Clients.objects.order_by('-id')
    say = Clients.objects.count()
    #STAT START
    osay = Orders.objects.count()
    csay = Clients.objects.count()
    psay = Products.objects.count()
    # STAT END
    user = request.user
    user_photo = user.photo
    first_name = user.first_name
    last_name = user.last_name
    email = user.email
    settings = Settings.objects.first()
    if request.user.is_superuser:
        vcheck = None
        echeck = None
    else:
        vcheck_instance = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=0).first()
        vcheck = vcheck_instance.item_id_id if vcheck_instance else None
        echeck_instance = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=1).first()
        echeck = echeck_instance.item_id_id if echeck_instance else None
        admin = ''
    if request.user.is_superuser:
        admin = Permissions.objects.all()

    perinfo = Perms.objects.filter(user_id=request.user.id)
    data = {'admin':admin,'perinfo':perinfo,'vcheck':vcheck,'echeck':echeck,'settings':settings,'info':info,'say':say,'delete1_id':id,'osay':osay,'csay':csay,'psay':psay,'user':user,'first_name':first_name,'last_name':last_name,'email':email,'user_photo':user_photo}
    return render(request,'clients.html',data)

@login_required
def udalit2(request,id,check=None):
    messages.info(request, 'Do you want to delete the department?',extra_tags='alert-info')
    info = Departments.objects.order_by('-id')
    say = Departments.objects.count()
             #STAT START
    dsay = Departments.objects.count()
    ssay = Staff.objects.count()
    psay = Positions.objects.count()
    # STAT END
    user = request.user
    user_photo = user.photo
    first_name = user.first_name
    last_name = user.last_name
    email = user.email
    settings = Settings.objects.first()
    if request.user.is_superuser:
        vcheck = None
        echeck = None
    else:
        vcheck_instance = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=0).first()
        vcheck = vcheck_instance.item_id_id if vcheck_instance else None
        echeck_instance = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=1).first()
        echeck = echeck_instance.item_id_id if echeck_instance else None
        admin = ''
    if request.user.is_superuser:
        admin = Permissions.objects.all()

    perinfo = Perms.objects.filter(user_id=request.user.id)
    data = {'admin':admin,'perinfo':perinfo,'vcheck':vcheck,'echeck':echeck,'settings':settings,'info':info,'say':say,'delete2_id':id,'dsay':dsay,'ssay':ssay,'psay':psay,'user':user,'first_name':first_name,'last_name':last_name,'email':email,'user_photo':user_photo}
    return render(request,'departments.html',data)

@login_required
def udalit3(request,id,check=None):
    messages.info(request, 'Do you want to delete the expense?',extra_tags='alert-info')
    info = Expense.objects.order_by('-id')
    say = Expense.objects.count()
             #STAT START
    esay = Orders.objects.count()
    # STAT END
    user = request.user
    user_photo = user.photo
    first_name = user.first_name
    last_name = user.last_name
    email = user.email
    settings = Settings.objects.first()
    if request.user.is_superuser:
        vcheck = None
        echeck = None
    else:
        vcheck_instance = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=0).first()
        vcheck = vcheck_instance.item_id_id if vcheck_instance else None
        echeck_instance = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=1).first()
        echeck = echeck_instance.item_id_id if echeck_instance else None
        admin = ''
    if request.user.is_superuser:
        admin = Permissions.objects.all()

    perinfo = Perms.objects.filter(user_id=request.user.id)

    data = {'admin':admin,'perinfo':perinfo,'vcheck':vcheck,'echeck':echeck,'settings':settings,'info':info,'say':say,'delete3_id':id,'esay':esay,'user':user,'first_name':first_name,'last_name':last_name,'email':email,'user_photo':user_photo}
    return render(request,'expense.html',data)

@login_required
def udalit4(request,id,check=None):
    messages.info(request, 'Do you want to delete the order?',extra_tags='alert-info')
    info = Orders.objects.order_by('-id')
    say = Orders.objects.count()
             #STAT START
    osay = Orders.objects.count()
    bsay = Brands.objects.count()
    csay = Clients.objects.count()
    psay = Products.objects.count()
    active_count = Orders.objects.filter(accept=0).count()
    completed_count = Orders.objects.filter(accept=1).count()
    # STAT END
    user = request.user
    user_photo = user.photo
    first_name = user.first_name
    last_name = user.last_name
    email = user.email
    settings = Settings.objects.first()
    if request.user.is_superuser:
        vcheck = None
        echeck = None
    else:
        vcheck_instance = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=0).first()
        vcheck = vcheck_instance.item_id_id if vcheck_instance else None
        echeck_instance = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=1).first()
        echeck = echeck_instance.item_id_id if echeck_instance else None
        admin = ''
    if request.user.is_superuser:
        admin = Permissions.objects.all()

    perinfo = Perms.objects.filter(user_id=request.user.id)

    data = {'active_count':active_count,'completed_count':completed_count,'admin':admin,'perinfo':perinfo,'vcheck':vcheck,'echeck':echeck,'settings':settings,'info':info,'say':say,'delete4_id':id,'osay':osay,'bsay':bsay,'csay':csay,'psay':psay,'user':user,'first_name':first_name,'last_name':last_name,'email':email,'user_photo':user_photo}
    return render(request,'orders.html',data)

@login_required
def udalit5(request,id,check=None):
    messages.info(request, 'Do you want to delete the task?',extra_tags='alert-info')
    info = Planner.objects.order_by('-id')
    say = Planner.objects.count()
             #STAT START
    osay = Orders.objects.count()
    bsay = Brands.objects.count()
    csay = Clients.objects.count()
    psay = Products.objects.count()
    active_count = Planner.objects.filter(accept=0).count()
    completed_count = Planner.objects.filter(accept=1).count()
    # STAT END
    user = request.user
    user_photo = user.photo
    first_name = user.first_name
    last_name = user.last_name
    email = user.email
    settings = Settings.objects.first()
    if request.user.is_superuser:
        vcheck = None
        echeck = None
    else:
        vcheck_instance = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=0).first()
        vcheck = vcheck_instance.item_id_id if vcheck_instance else None
        echeck_instance = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=1).first()
        echeck = echeck_instance.item_id_id if echeck_instance else None
        admin = ''
    if request.user.is_superuser:
        admin = Permissions.objects.all()

    perinfo = Perms.objects.filter(user_id=request.user.id)

    data = {'active_count':active_count,'completed_count':completed_count,'admin':admin,'perinfo':perinfo,'vcheck':vcheck,'echeck':echeck,'settings':settings,'info':info,'say':say,'delete5_id':id,'osay':osay,'bsay':bsay,'csay':csay,'psay':psay,'user':user,'first_name':first_name,'last_name':last_name,'email':email,'user_photo':user_photo}
    return render(request,'planner.html',data)

@login_required
def udalit6(request,id,check=None):
    messages.info(request, 'Do you want to delete the position?',extra_tags='alert-info')
    info = Positions.objects.order_by('-id')
    say = Positions.objects.count()
             #STAT START
    dsay = Departments.objects.count()
    psay = Positions.objects.count()
    ssay = Staff.objects.count()
    # STAT END
    user = request.user
    user_photo = user.photo
    first_name = user.first_name
    last_name = user.last_name
    email = user.email
    settings = Settings.objects.first()

    settings = Settings.objects.first()
    if request.user.is_superuser:
        vcheck = None
        echeck = None
    else:
        vcheck_instance = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=0).first()
        vcheck = vcheck_instance.item_id_id if vcheck_instance else None
        echeck_instance = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=1).first()
        echeck = echeck_instance.item_id_id if echeck_instance else None
        admin = ''
    if request.user.is_superuser:
        admin = Permissions.objects.all()

    perinfo = Perms.objects.filter(user_id=request.user.id)
    data = {'admin':admin,'perinfo':perinfo,'vcheck':vcheck,'echeck':echeck,'settings':settings,'info':info,'say':say,'delete6_id':id,'ssay':ssay,'dsay':dsay,'psay':psay,'user':user,'first_name':first_name,'last_name':last_name,'email':email,'user_photo':user_photo}
    return render(request,'positions.html',data)

@login_required
def udalit7(request,id,check=None):
    messages.info(request, 'Do you want to delete the product?',extra_tags='alert-info')
    info = Products.objects.order_by('-id')
    say = Products.objects.count()
             #STAT START
    osay = Orders.objects.count()
    bsay = Brands.objects.count()
    csay = Clients.objects.count()
    psay = Products.objects.count()
    # STAT END
    user = request.user
    user_photo = user.photo
    first_name = user.first_name
    last_name = user.last_name
    email = user.email
    settings = Settings.objects.first()
    if request.user.is_superuser:
        vcheck = None
        echeck = None
    else:
        vcheck_instance = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=0).first()
        vcheck = vcheck_instance.item_id_id if vcheck_instance else None
        echeck_instance = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=1).first()
        echeck = echeck_instance.item_id_id if echeck_instance else None
        admin = ''
    if request.user.is_superuser:
        admin = Permissions.objects.all()
    perinfo = Perms.objects.filter(user_id=request.user.id)

    data = {'admin':admin,'perinfo':perinfo,'vcheck':vcheck,'echeck':echeck,'settings':settings,'info':info,'say':say,'delete7_id':id,'osay':osay,'bsay':bsay,'csay':csay,'psay':psay,'user':user,'first_name':first_name,'last_name':last_name,'email':email,'user_photo':user_photo}
    return render(request,'products.html',data)

@login_required
def udalit8(request,id,check=None):
    messages.info(request, 'Do you want to delete the staff?',extra_tags='alert-info')
    info = Staff.objects.order_by('-id')
    say = Staff.objects.count()
             #STAT START
    dsay = Departments.objects.count()
    ssay = Staff.objects.count()
    psay = Products.objects.count()
    # STAT END
    user = request.user
    user_photo = user.photo
    first_name = user.first_name
    last_name = user.last_name
    email = user.email
    settings = Settings.objects.first()
    if request.user.is_superuser:
        vcheck = None
        echeck = None
    else:
        vcheck_instance = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=0).first()
        vcheck = vcheck_instance.item_id_id if vcheck_instance else None
        echeck_instance = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=1).first()
        echeck = echeck_instance.item_id_id if echeck_instance else None
        admin = ''
    if request.user.is_superuser:
        admin = Permissions.objects.all()

    perinfo = Perms.objects.filter(user_id=request.user.id)

    data = {'admin':admin,'perinfo':perinfo,'vcheck':vcheck,'echeck':echeck,'settings':settings,'info':info,'say':say,'delete8_id':id,'dsay':dsay,'ssay':ssay,'psay':psay,'user':user,'first_name':first_name,'last_name':last_name,'email':email,'user_photo':user_photo}
    return render(request,'staff.html',data)

@login_required
def udalit9(request,id,check=None):
    messages.info(request, 'Do you want to delete the supplier?',extra_tags='alert-info')
    info = Suppliers.objects.order_by('-id')
    say = Suppliers.objects.count()
             #STAT START
    osay = Orders.objects.count()
    bsay = Brands.objects.count()
    csay = Clients.objects.count()
    psay = Products.objects.count()
    # STAT END
    user = request.user
    user_photo = user.photo
    first_name = user.first_name
    last_name = user.last_name
    email = user.email
    settings = Settings.objects.first()
    if request.user.is_superuser:
        vcheck = None
        echeck = None
    else:
        vcheck_instance = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=0).first()
        vcheck = vcheck_instance.item_id_id if vcheck_instance else None
        echeck_instance = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=1).first()
        echeck = echeck_instance.item_id_id if echeck_instance else None
        admin = ''
    if request.user.is_superuser:
        admin = Permissions.objects.all()

    perinfo = Perms.objects.filter(user_id=request.user.id)

    data = {'admin':admin,'perinfo':perinfo,'vcheck':vcheck,'echeck':echeck,'settings':settings,'info':info,'say':say,'delete9_id':id,'osay':osay,'bsay':bsay,'csay':csay,'psay':psay,'user':user,'first_name':first_name,'last_name':last_name,'email':email,'user_photo':user_photo}
    return render(request,'suppliers.html',data)
#UDALIT FUNCTION END

#EDIT FUNCTIONS START
@login_required
def edit(request, id, check=None, user_id=None):
    info = Brands.objects.order_by('-id')
    say = Brands.objects.count()

    # STAT START
    osay = Orders.objects.count()
    bsay = Brands.objects.count()
    psay = Products.objects.count()
    # STAT END

    edit = Brands.objects.get(id=id)
    user = request.user
    user_id = request.user.id
    first_name = user.first_name
    last_name = user.last_name
    email = user.email
    user_photo = user.photo
    settings = Settings.objects.first()
    perinfo = Perms.objects.filter(user_id=request.user.id)
    if request.user.is_superuser:
        vcheck = None
        echeck = None
    else:
        vcheck = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=0).count()
        echeck = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=1).count()
    admin = ''
    if request.user.is_superuser:
        admin = Permissions.objects.all()
    data = {'user_id': user_id, 'admin': admin, 'perinfo': perinfo, 'vcheck': vcheck, 'echeck': echeck, 'settings': settings, 'info': info, 'say': say, 'edit': edit, 'osay': osay, 'bsay': bsay, 'psay': psay, 'osay': osay, 'user': user, 'first_name': first_name, 'last_name': last_name, 'email': email, 'user_photo': user_photo}
    if 'cancel' in request.POST:
        return HttpResponseRedirect(reverse('brands'))
    return render(request, 'brands.html', data)



    
@login_required
def edit1(request, id, check=None):
    all_clients = Clients.objects.order_by('-id')
    client_count = Clients.objects.count()
    client = Clients.objects.get(id=id)
    #STAT START
    osay = Orders.objects.count()
    csay = Clients.objects.count()
    psay = Products.objects.count()
    # STAT END
    user = request.user
    user_photo = user.photo
    first_name = user.first_name
    last_name = user.last_name
    email = user.email

    if request.method == 'POST':
        # Получите идентификатор клиента из POST-запроса
        id = request.POST.get('id')
        client = Clients.objects.get(id=id)

        new_name = request.POST.get('name', '')
        new_surname = request.POST.get('surname', '')
        new_email = request.POST.get('email', '')
        new_telephone = request.POST.get('telephone', '')
        new_company = request.POST.get('company', '')

        if 'cancel' in request.POST:
            return HttpResponseRedirect(reverse('clients'))

        # Check if a new photo was uploaded
        if 'new_photo' in request.FILES:
            upload = request.FILES['new_photo']
            fs = FileSystemStorage()
            file = fs.save(upload.name, upload)
            file_url = fs.url(file)

            # Update the client's photo
            client.photo = file_url

        if new_name != '' and new_surname != '' and new_email != '' and new_telephone != '' and new_company != '':
            client.name = new_name
            client.surname = new_surname
            client.email = new_email
            client.telephone = new_telephone
            client.company = new_company
            client.save()
            messages.info(request, 'Client has been updated successfully!',extra_tags='alert-success')
            return HttpResponseRedirect(reverse('clients'))

    settings = Settings.objects.first()
    if request.user.is_superuser:
        vcheck = None
        echeck = None
    else:
        vcheck_instance = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=0).first()
        vcheck = vcheck_instance.item_id_id if vcheck_instance else None
        echeck_instance = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=1).first()
        echeck = echeck_instance.item_id_id if echeck_instance else None
        admin = ''
    if request.user.is_superuser:
        admin = Permissions.objects.all()

    perinfo = Perms.objects.filter(user_id=request.user.id)
    data = {'admin':admin,'perinfo':perinfo,'vcheck':vcheck,'echeck':echeck,'settings':settings,'info': all_clients, 'say': client_count, 'edit1': client,'osay':osay,'csay':csay,'psay':psay,'user':user,'first_name':first_name,'last_name':last_name,'email':email,'user_photo':user_photo}

    return render(request, 'clients.html', data)

@login_required
def edit2(request,id,check=None):

    info = Departments.objects.order_by('-id')
    say = Departments.objects.count()
    edit2 = Departments.objects.get(id=id)
    #STAT START
    dsay = Departments.objects.count()
    ssay = Staff.objects.count()
    psay = Positions.objects.count()
    # STAT END
    user = request.user
    first_name = user.first_name
    last_name = user.last_name
    email = user.email
    user_photo = user.photo
    settings = Settings.objects.first()
    if request.user.is_superuser:
        vcheck = None
        echeck = None
    else:
        vcheck_instance = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=0).first()
        vcheck = vcheck_instance.item_id_id if vcheck_instance else None
        echeck_instance = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=1).first()
        echeck = echeck_instance.item_id_id if echeck_instance else None
        admin = ''
    if request.user.is_superuser:
        admin = Permissions.objects.all()

    perinfo = Perms.objects.filter(user_id=request.user.id)
    data = {'admin':admin,'perinfo':perinfo,'vcheck':vcheck,'echeck':echeck,'settings':settings,'info':info,'say':say,'edit2':edit2,'dsay':dsay,'ssay':ssay,'psay':psay,'user':user,'first_name':first_name,'last_name':last_name,'email':email,'user_photo':user_photo}

    return render(request,'departments.html',data)

@login_required
def edit3(request,id,check=None):

    info = Expense.objects.order_by('-id')
    say = Expense.objects.count()
    edit3 = Expense.objects.get(id=id)
    #STAT START
    esay = Expense.objects.count()
    # STAT 
    user = request.user
    first_name = user.first_name
    last_name = user.last_name
    email = user.email
    user_photo = user.photo
    settings = Settings.objects.first()
    if request.user.is_superuser:
        vcheck = None
        echeck = None
    else:
        vcheck_instance = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=0).first()
        vcheck = vcheck_instance.item_id_id if vcheck_instance else None
        echeck_instance = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=1).first()
        echeck = echeck_instance.item_id_id if echeck_instance else None
        admin = ''
    if request.user.is_superuser:
        admin = Permissions.objects.all()

    perinfo = Perms.objects.filter(user_id=request.user.id)

    data = {'admin':admin,'perinfo':perinfo,'vcheck':vcheck,'echeck':echeck,'settings':settings,'info':info,'say':say,'edit3':edit3,'esay':esay,'user':user,'first_name':first_name,'last_name':last_name,'email':email,'user_photo':user_photo}

    return render(request,'expense.html',data)

@login_required
def edit4(request, id, check=None):
    info = Orders.objects.order_by('-id')
    say = Orders.objects.count()
    #STAT START
    osay = Orders.objects.count()
    bsay = Brands.objects.count()
    csay = Clients.objects.count()
    psay = Products.objects.count()
    # STAT END
    user = request.user
    first_name = user.first_name
    last_name = user.last_name
    email = user.email
    user_photo = user.photo
    try:
        edit4 = Orders.objects.get(id=id)
    except Orders.DoesNotExist:
        messages.info(request, 'Order does not exist!',extra_tags='alert-danger')

    cinfo = Clients.objects.all()
    pinfo = Products.objects.all()
    active_count = Planner.objects.filter(accept=0).count()
    completed_count = Planner.objects.filter(accept=1).count()
    settings = Settings.objects.first()
    if request.user.is_superuser:
        vcheck = None
        echeck = None
    else:
        vcheck_instance = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=0).first()
        vcheck = vcheck_instance.item_id_id if vcheck_instance else None
        echeck_instance = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=1).first()
        echeck = echeck_instance.item_id_id if echeck_instance else None
        admin = ''
    if request.user.is_superuser:
        admin = Permissions.objects.all()
    perinfo = Perms.objects.filter(user_id=request.user.id)

    data = {'admin':admin,'perinfo':perinfo,'vcheck':vcheck,'echeck':echeck,'settings':settings,'info': info, 'say': say, 'edit4': edit4, 'cinfo': cinfo, 'pinfo': pinfo, 'active_count': active_count, 'completed_count': completed_count,'osay':osay,'bsay':bsay,'csay':csay,'psay':psay,'user':user,'first_name':first_name,'last_name':last_name,'email':email,'user_photo':user_photo}
    return render(request, 'orders.html', data)

@login_required
def edit5(request, id,check=None):
    info = Planner.objects.order_by('-id')
    say = Planner.objects.count()
    active_count = Planner.objects.filter(accept=0).count()
    completed_count = Planner.objects.filter(accept=1).count()
    #STAT START
    osay = Orders.objects.count()
    bsay = Brands.objects.count()
    csay = Clients.objects.count()
    psay = Products.objects.count()
    # STAT END
    user = request.user
    first_name = user.first_name
    last_name = user.last_name
    email = user.email
    user_photo = user.photo

    try:
        edit5 = Planner.objects.get(id=id)
    except Planner.DoesNotExist:
        messages.info(request, 'Planner does not exist!',extra_tags='alert-danger')

    sinfo = Staff.objects.all()
    active_count = Planner.objects.filter(accept=0).count()
    completed_count = Planner.objects.filter(accept=1).count()
    settings = Settings.objects.first()
    if request.user.is_superuser:
        vcheck = None
        echeck = None
    else:
        vcheck_instance = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=0).first()
        vcheck = vcheck_instance.item_id_id if vcheck_instance else None
        echeck_instance = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=1).first()
        echeck = echeck_instance.item_id_id if echeck_instance else None
        admin = ''
    if request.user.is_superuser:
        admin = Permissions.objects.all()

    perinfo = Perms.objects.filter(user_id=request.user.id)


    data = {
        'admin':admin,
        'perinfo':perinfo,
        'vcheck':vcheck,
        'echeck':echeck,
        'settings':settings,
        'info': info,
        'say': say,
        'edit5': edit5,
        'sinfo': sinfo,
        'active_count': active_count,
        'completed_count': completed_count,
        'osay':osay,
        'psay':psay,
        'bsay':bsay,
        'csay':csay,
        'user':user,
        'first_name':first_name,
        'last_name':last_name,
        'email':email,
        'user_photo':user_photo
     
    }

    return render(request, 'planner.html', data)



@login_required
def edit6(request,id,check=None):

    info = Positions.objects.order_by('-id')
    say = Positions.objects.count()
    dinfo = Departments.objects.all()
    edit6 = Positions.objects.get(id=id)
    #STAT START
    dsay = Departments.objects.count()
    psay = Positions.objects.count()
    ssay = Staff.objects.count()
    # STAT END
    user = request.user
    first_name = user.first_name
    last_name = user.last_name
    email = user.email
    user_photo = user.photo
    settings = Settings.objects.first()
    if request.user.is_superuser:
        vcheck = None
        echeck = None
    else:
        vcheck_instance = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=0).first()
        vcheck = vcheck_instance.item_id_id if vcheck_instance else None
        echeck_instance = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=1).first()
        echeck = echeck_instance.item_id_id if echeck_instance else None
        admin = ''
    if request.user.is_superuser:
        admin = Permissions.objects.all()

    perinfo = Perms.objects.filter(user_id=request.user.id)
    data = {'admin':admin,'perinfo':perinfo,'vcheck':vcheck,'echeck':echeck,'settings':settings,'info':info,'say':say,'edit6':edit6,'dinfo':dinfo,'dsay':dsay,'ssay':ssay,'psay':psay,'user':user,'first_name':first_name,'last_name':last_name,'email':email,'user_photo':user_photo}

    return render(request,'positions.html',data)

@login_required
def edit7(request, id,check=None):
    info = Products.objects.order_by('-id')
    say = Products.objects.count()
    edit7 = Products.objects.get(id=id)
    binfo = Brands.objects.all()
    sinfo = Suppliers.objects.all()
        #STAT START
    osay = Orders.objects.count()
    bsay = Brands.objects.count()
    csay = Clients.objects.count()
    psay = Products.objects.count()
    # STAT END
    user = request.user
    first_name = user.first_name
    last_name = user.last_name
    email = user.email
    user_photo = user.photo
    if 'cancel' in request.POST:
        return HttpResponseRedirect(reverse('products'))

    if 'new_photo' in request.FILES:
        upload = request.FILES['new_photo']
        fs = FileSystemStorage()
        file = fs.save(upload.name, upload)
        file_url = fs.url(file)

        edit7.photo = file_url

    settings = Settings.objects.first()
    if request.user.is_superuser:
        vcheck = None
        echeck = None
    else:
        vcheck_instance = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=0).first()
        vcheck = vcheck_instance.item_id_id if vcheck_instance else None
        echeck_instance = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=1).first()
        echeck = echeck_instance.item_id_id if echeck_instance else None
        admin = ''
    if request.user.is_superuser:
        admin = Permissions.objects.all()

    perinfo = Perms.objects.filter(user_id=request.user.id)

    data = {'admin':admin,'perinfo':perinfo,'vcheck':vcheck,'echeck':echeck,'settings':settings,'info': info, 'say': say, 'edit7': edit7, 'binfo': binfo,'sinfo':sinfo,'osay':osay,'bsay':bsay,'csay':csay,'psay':psay,'user':user,'first_name':first_name,'last_name':last_name,'email':email,'user_photo':user_photo}  
    return render(request, 'products.html', data)


@login_required
def edit8(request, id,check=None):
    info = Staff.objects.order_by('-id')
    say = Staff.objects.count()
    edit8 = Staff.objects.get(id=id)
    pinfo = Positions.objects.all()
        #STAT START
    dsay = Departments.objects.count()
    ssay = Staff.objects.count()
    psay = Products.objects.count()
    # STAT END
    user = request.user
    first_name = user.first_name
    last_name = user.last_name
    email = user.email
    user_photo = user.photo
    if request.method == 'POST':
        new_name = request.POST.get('name', '')
        new_email = request.POST.get('email', '')
        new_phone = request.POST.get('phone', '')
        new_position_id = request.POST.get('position_id', '')  # Добавляем обработку позиции

        if 'cancel' in request.POST:
            return HttpResponseRedirect(reverse('staff'))

        if 'new_photo' in request.FILES:
            upload = request.FILES['new_photo']
            fs = FileSystemStorage()
            file = fs.save(upload.name, upload)
            file_url = fs.url(file)
            edit8.photo = file_url

        if new_name != '' and new_email != '' and new_phone != '':
            edit8.name = new_name
            edit8.email = new_email
            edit8.phone = new_phone

            if new_position_id:
                edit8.position_id = Positions.objects.get(id=new_position_id)  # Сохраняем новую позицию

            edit8.save()
            messages.info(request, 'Staff member has been updated successfully!',extra_tags='alert-success')
            return HttpResponseRedirect(reverse('staff'))
    settings = Settings.objects.first()
    if request.user.is_superuser:
        vcheck = None
        echeck = None
    else:
        vcheck_instance = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=0).first()
        vcheck = vcheck_instance.item_id_id if vcheck_instance else None
        echeck_instance = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=1).first()
        echeck = echeck_instance.item_id_id if echeck_instance else None
        admin = ''
    if request.user.is_superuser:
        admin = Permissions.objects.all()
    perinfo = Perms.objects.filter(user_id=request.user.id)

    data = {'admin':admin,'perinfo':perinfo,'vcheck':vcheck,'echeck':echeck,'settings':settings,'info': info, 'say': say, 'edit8': edit8, 'pinfo': pinfo,'dsay':dsay,'psay':psay,'ssay':ssay,'user':user,'first_name':first_name,'last_name':last_name,'email':email,'user_photo':user_photo}

    return render(request, 'staff.html', data)


@login_required
def edit9(request, id,check=None):
    info = Suppliers.objects.order_by('-id')
    say = Suppliers.objects.count()
    edit9 = Suppliers.objects.get(id=id)
    #STAT START
    osay = Orders.objects.count()
    bsay = Brands.objects.count()
    csay = Clients.objects.count()
    psay = Products.objects.count()
    # STAT END
    user = request.user
    first_name = user.first_name
    last_name = user.last_name
    email = user.email
    user_photo = user.photo
    if request.method == 'POST':
        new_name = request.POST.get('name', '')
        new_email = request.POST.get('email', '')
        new_phone = request.POST.get('phone', '')

        if 'cancel' in request.POST:
            return HttpResponseRedirect(reverse('suppliers'))

        if 'new_photo' in request.FILES:
            upload = request.FILES['new_photo']
            fs = FileSystemStorage()
            file = fs.save(upload.name, upload)
            file_url = fs.url(file)

            edit9.photo = file_url

        if new_name != '' and new_email != '' and new_phone != '':
            edit9.name = new_name
            edit9.email = new_email
            edit9.phone = new_phone
            edit9.save()
            messages.info(request, 'Supplier has been updated successfully!',extra_tags='alert-success')
            return HttpResponseRedirect(reverse('suppliers'))
    
    settings = Settings.objects.first()
    if request.user.is_superuser:
        vcheck = None
        echeck = None
    else:
        vcheck_instance = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=0).first()
        vcheck = vcheck_instance.item_id_id if vcheck_instance else None
        echeck_instance = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=1).first()
        echeck = echeck_instance.item_id_id if echeck_instance else None
        admin = ''
    if request.user.is_superuser:
        admin = Permissions.objects.all()

    perinfo = Perms.objects.filter(user_id=request.user.id)
    data = {'admin':admin,'perinfo':perinfo,'vcheck':vcheck,'echeck':echeck,'settings':settings,'info': info, 'say': say, 'edit9': edit9,'osay':osay,'bsay':bsay,'csay':csay,'psay':psay,'user':user,'first_name':first_name,'last_name':last_name,'email':email,'user_photo':user_photo}

    return render(request, 'suppliers.html', data)


#EDIT FUNCTIONS END

#UPDATE FUNCTION START
@login_required
def update(request, id, check=None,user_id=None):
    info = Brands.objects.get(id=id)
    x = request.POST.get('brand', '')
    user_id=request.user.id
    echeck = None
    vcheck = None


    if 'new_photo' in request.FILES:
        upload = request.FILES['new_photo']
        fs = FileSystemStorage()
        file = fs.save(upload.name, upload)
        file_url = fs.url(file)
        info.photo = file_url

    info.brand = x
    info.save()

    if not request.user.is_superuser:
        vcheck = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=0).count()
        echeck = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=1).count()
    admin = ''
    if request.user.is_superuser:
        admin = Permissions.objects.all()

    messages.info(request, 'Brand has been updated successfully!',extra_tags='alert-success')
    data = {'echeck': echeck, 'vcheck': vcheck, 'user_id': user_id, 'admin': admin, 'info': info}
    if 'cancel' in request.POST:
        return HttpResponseRedirect(reverse('brands'), {'data': data})
    if request.user.is_superuser:
        return HttpResponseRedirect(reverse('brands'))  # Redirect to 'brands' for superuser
    else:
        return HttpResponseRedirect(reverse('brands', args=[check, id]))  # Redirect to 'brands/check/id' for non-superuser




@login_required
def update1(request, id, check=None,user_id=None):
    client = Clients.objects.get(id=id)

    if 'cancel' in request.POST:
        return HttpResponseRedirect(reverse('clients'))

    new_name = request.POST['name']
    new_surname = request.POST['surname']
    new_email = request.POST['email']
    new_telephone = request.POST['telephone']
    new_company = request.POST['company']

    # Check if a new photo was uploaded
    if 'new_photo' in request.FILES:
        upload = request.FILES['new_photo']
        fs = FileSystemStorage()
        file = fs.save(upload.name, upload)
        file_url = fs.url(file)

        # Update the client's photo
        client.photo = file_url

    client.name = new_name
    client.surname = new_surname
    client.email = new_email
    client.telephone = new_telephone
    client.company = new_company

    client.save()

    if request.user.is_superuser:
        vcheck = None
        echeck = None
    else:
        vcheck_instance = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=0).first()
        vcheck = vcheck_instance.item_id_id if vcheck_instance else None
        echeck_instance = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=1).first()
        echeck = echeck_instance.item_id_id if echeck_instance else None
        admin = ''
    if request.user.is_superuser:
        admin = Permissions.objects.all()

    messages.info(request, 'Client has been updated successfully!',extra_tags='alert-success')
    data = {'echeck': echeck, 'vcheck': vcheck, 'user_id': user_id, 'admin': admin}
    if 'cancel' in request.POST:
        return HttpResponseRedirect(reverse('clients'), {'data': data})
    if request.user.is_superuser:
        return HttpResponseRedirect(reverse('clients'))  # Redirect to 'brands' for superuser
    else:
        return HttpResponseRedirect(reverse('clients', args=[check, id])) 

@login_required
def update2(request, id, check=None , user_id=None):
    info = Departments.objects.get(id=id)
    x = request.POST['department']

    if 'cancel' in request.POST:
        return HttpResponseRedirect(reverse('departments'))
    
    info.department = x
    info.save()

    if request.user.is_superuser:
        vcheck = None
        echeck = None
    else:
        vcheck_instance = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=0).first()
        vcheck = vcheck_instance.item_id_id if vcheck_instance else None
        echeck_instance = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=1).first()
        echeck = echeck_instance.item_id_id if echeck_instance else None
        admin = ''
    if request.user.is_superuser:
        admin = Permissions.objects.all()

    messages.info(request, 'Department has been updated successfully!',extra_tags='alert-success')
    data = {'echeck': echeck, 'vcheck': vcheck, 'user_id': user_id, 'admin': admin}
    if 'cancel' in request.POST:
        return HttpResponseRedirect(reverse('departments'), {'data': data})
    if request.user.is_superuser:
        return HttpResponseRedirect(reverse('departments'))  # Redirect to 'brands' for superuser
    else:
        return HttpResponseRedirect(reverse('departments', args=[check, id])) 


@login_required
def update3(request, id, check=None, user_id=None):
    info = Expense.objects.get(id=id)
    x = request.POST['appointment']
    y = request.POST['amount']

    if 'cancel' in request.POST:
        return HttpResponseRedirect(reverse('expense'))
    
    info.appointment = x
    info.amount = y
    info.save()

    if request.user.is_superuser:
        vcheck = None
        echeck = None
    else:
        vcheck_instance = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=0).first()
        vcheck = vcheck_instance.item_id_id if vcheck_instance else None
        echeck_instance = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=1).first()
        echeck = echeck_instance.item_id_id if echeck_instance else None
        admin = ''
    if request.user.is_superuser:
        admin = Permissions.objects.all()

    messages.info(request, 'Expense has been updated successfully!',extra_tags='alert-success')
    data = {'echeck': echeck, 'vcheck': vcheck, 'user_id': user_id, 'admin': admin}
    if 'cancel' in request.POST:
        return HttpResponseRedirect(reverse('expense'), {'data': data})
    if request.user.is_superuser:
        return HttpResponseRedirect(reverse('expense'))  # Redirect to 'brands' for superuser
    else:
        return HttpResponseRedirect(reverse('expense', args=[check, id])) 

@login_required
def update4(request,id, check=None, user_id=None):
    if request.method == 'POST':
        info = Orders.objects.get(id=id)
        client_id = request.POST.get('client_id')
        product_id = request.POST.get('product_id')
        x = request.POST.get('quantity')

        if 'cancel' in request.POST:
            return HttpResponseRedirect(reverse('orders'))

        if x:
            info.quantity = x

            if client_id:
                info.client_id = Clients.objects.get(id=client_id)
            
            if product_id:
                info.product_id = Products.objects.get(id=product_id)

            info.save()

            if request.user.is_superuser:
                vcheck = None
                echeck = None
            else:
                vcheck_instance = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=0).first()
                vcheck = vcheck_instance.item_id_id if vcheck_instance else None
                echeck_instance = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=1).first()
                echeck = echeck_instance.item_id_id if echeck_instance else None
                admin = ''
            if request.user.is_superuser:
                admin = Permissions.objects.all()

            messages.info(request, 'Order has been updated successfully!',extra_tags='alert-success')
            data = {'echeck': echeck, 'vcheck': vcheck, 'user_id': user_id, 'admin': admin}
            if 'cancel' in request.POST:
                return HttpResponseRedirect(reverse('orders'), {'data': data})
            if request.user.is_superuser:
                return HttpResponseRedirect(reverse('orders'))  # Redirect to 'brands' for superuser
            else:
                return HttpResponseRedirect(reverse('orders', args=[check, id])) 



@login_required
def update5(request,id, check=None ,user_id=None):
    info = Planner.objects.get(id=id)
    x = request.POST.get('task', '')
    y = request.POST.get('dtask', '')
    z = request.POST.get('ttask', '')
    staff_id = request.POST.get('staff_id')


    if 'cancel' in request.POST:
        return HttpResponseRedirect(reverse('planner'))
    
    info.task = x
    info.dtask = y
    info.ttask = z

    if staff_id:
        info.staff_id = Staff.objects.get(id=staff_id)

    info.save()
    messages.info(request, 'Task has been updated successfully!',extra_tags='alert-success')
    if request.user.is_superuser:
        vcheck = None
        echeck = None
    else:
        vcheck_instance = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=0).first()
        vcheck = vcheck_instance.item_id_id if vcheck_instance else None
        echeck_instance = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=1).first()
        echeck = echeck_instance.item_id_id if echeck_instance else None
        admin = ''
    if request.user.is_superuser:
        admin = Permissions.objects.all()

    active_count = Planner.objects.filter(accept=0).count()
    completed_count = Planner.objects.filter(accept=1).count()

    data = {'active_count':active_count,'completed_count':completed_count,'echeck': echeck, 'vcheck': vcheck, 'user_id': user_id, 'admin': admin}
    if 'cancel' in request.POST:
        return HttpResponseRedirect(reverse('planner'), {'data': data})
    if request.user.is_superuser:
        return HttpResponseRedirect(reverse('planner'))  # Redirect to 'brands' for superuser
    else:
        return HttpResponseRedirect(reverse('planner', args=[check, id])) 

@login_required
def update6(request,id, check=None , user_id=None):
    info = Positions.objects.get(id=id)
    x = request.POST['positions']
    department_id = request.POST.get('department_id')

    if 'cancel' in request.POST:
        return HttpResponseRedirect(reverse('positions'))
    
    info.positions = x

    if department_id:
        info.department_id = Departments.objects.get(id=department_id)

    info.save()
    messages.info(request, 'Position has been updated successfully!',extra_tags='alert-success')
    if request.user.is_superuser:
        vcheck = None
        echeck = None
    else:
        vcheck_instance = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=0).first()
        vcheck = vcheck_instance.item_id_id if vcheck_instance else None
        echeck_instance = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=1).first()
        echeck = echeck_instance.item_id_id if echeck_instance else None
        admin = ''
    if request.user.is_superuser:
        admin = Permissions.objects.all()


    data = {'echeck': echeck, 'vcheck': vcheck, 'user_id': user_id, 'admin': admin}
    if 'cancel' in request.POST:
        return HttpResponseRedirect(reverse('positions'), {'data': data})
    if request.user.is_superuser:
        return HttpResponseRedirect(reverse('positions'))  # Redirect to 'brands' for superuser
    else:
        return HttpResponseRedirect(reverse('positions', args=[check, id])) 

@login_required
def update7(request,id, check=None , user_id=None):
    info = Products.objects.get(id=id)
    x = request.POST.get('name', '')
    y = request.POST.get('purchase', '0')  # Default value is '0'
    z = request.POST.get('sale', '0')      # Default value is '0'
    a = request.POST.get('quantity', '0')  # Default value is '0'
    brand_id = request.POST['brand_id']
    supplier_id = request.POST['supplier_id']

    if 'cancel' in request.POST:
        return HttpResponseRedirect(reverse('products'))

    # Check if a new photo was uploaded
    if 'new_photo' in request.FILES:
        upload = request.FILES['new_photo']
        fs = FileSystemStorage()
        file = fs.save(upload.name, upload)
        file_url = fs.url(file)

        # Update the staff member's photo
        info.photo = file_url

    info.name = x
    info.purchase = float(y)  # Convert 'y' to a floating-point number
    info.sale = float(z)      # Convert 'z' to a floating-point number
    info.quantity = int(a)    # Convert 'a' to an integer

    if brand_id:
        info.brand_id = Brands.objects.get(id=brand_id)
    
    if supplier_id:
        info.supplier_id = Suppliers.objects.get(id=supplier_id)

    info.save()
    if request.user.is_superuser:
        vcheck = None
        echeck = None
    else:
        vcheck_instance = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=0).first()
        vcheck = vcheck_instance.item_id_id if vcheck_instance else None
        echeck_instance = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=1).first()
        echeck = echeck_instance.item_id_id if echeck_instance else None
        admin = ''
    if request.user.is_superuser:
        admin = Permissions.objects.all()


    data = {'echeck': echeck, 'vcheck': vcheck, 'user_id': user_id, 'admin': admin}
    messages.info(request, 'Product has been updated successfully!',extra_tags='alert-success')
    if 'cancel' in request.POST:
        return HttpResponseRedirect(reverse('products'), {'data': data})
    if request.user.is_superuser:
        return HttpResponseRedirect(reverse('products'))  # Redirect to 'brands' for superuser
    else:
        return HttpResponseRedirect(reverse('products', args=[check, id])) 


@login_required
def update8(request,id, check=None ,user_id=None):
    info = Staff.objects.get(id=id)
    x = request.POST['name']
    y = request.POST['surname']
    z = request.POST['email']
    a = request.POST['telephone']
    b = request.POST['salary']
    c = request.POST['work']
    position_id = request.POST['position_id']

    if 'cancel' in request.POST:
        return HttpResponseRedirect(reverse('staff'))

    # Check if a new photo was uploaded
    if 'new_photo' in request.FILES:
        upload = request.FILES['new_photo']
        fs = FileSystemStorage()
        file = fs.save(upload.name, upload)
        file_url = fs.url(file)

        # Update the staff member's photo
        info.photo = file_url

    info.name = x
    info.surname = y
    info.email = z
    info.telephone = a
    info.salary = b
    info.work = c

    if position_id:
        info.position_id = Positions.objects.get(id=position_id)

    info.save()

    if request.user.is_superuser:
        vcheck = None
        echeck = None
    else:
        vcheck_instance = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=0).first()
        vcheck = vcheck_instance.item_id_id if vcheck_instance else None
        echeck_instance = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=1).first()
        echeck = echeck_instance.item_id_id if echeck_instance else None
        admin = ''
    if request.user.is_superuser:
        admin = Permissions.objects.all()

    messages.info(request, 'Staff has been updated successfully!',extra_tags='alert-success')

    data = {'echeck': echeck, 'vcheck': vcheck, 'user_id': user_id, 'admin': admin}
    if 'cancel' in request.POST:
        return HttpResponseRedirect(reverse('staff'), {'data': data})
    if request.user.is_superuser:
        return HttpResponseRedirect(reverse('staff'))  # Redirect to 'brands' for superuser
    else:
        return HttpResponseRedirect(reverse('staff', args=[check, id])) 

@login_required
def update9(request,id,check=None,user_id=None):
    info = Suppliers.objects.get(id=id)
    c = request.POST.get('firm', '')
    x = request.POST.get('name', '')
    y = request.POST.get('surname', '')
    z = request.POST.get('email', '')
    a = request.POST.get('telephone', '')
    b = request.POST.get('address', '')

    if 'cancel' in request.POST:
        return HttpResponseRedirect(reverse('suppliers'))

    # Check if a new photo was uploaded
    if 'new_photo' in request.FILES:
        upload = request.FILES['new_photo']
        fs = FileSystemStorage()
        file = fs.save(upload.name, upload)
        file_url = fs.url(file)

        # Update the supplier's photo
        info.photo = file_url

    info.firm = c
    info.name = x
    info.surname = y
    info.email = z
    info.telephone = a
    info.address = b
    info.save()

    if request.user.is_superuser:
        vcheck = None
        echeck = None
    else:
        vcheck_instance = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=0).first()
        vcheck = vcheck_instance.item_id_id if vcheck_instance else None
        echeck_instance = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=1).first()
        echeck = echeck_instance.item_id_id if echeck_instance else None
        admin = ''
    if request.user.is_superuser:
        admin = Permissions.objects.all()

    messages.info(request, 'Supplier has been updated successfully!',extra_tags='alert-success')

    data = {'echeck': echeck, 'vcheck': vcheck, 'user_id': user_id, 'admin': admin}
    if 'cancel' in request.POST:
        return HttpResponseRedirect(reverse('suppliers'), {'data': data})
    if request.user.is_superuser:
        return HttpResponseRedirect(reverse('suppliers'))  # Redirect to 'brands' for superuser
    else:
        return HttpResponseRedirect(reverse('suppliers', args=[check, id])) 



#UPDATE FUNCTIONS END
@login_required
def accept_order(request, order_id, check=None):
    settings = Settings.objects.first()
    if request.user.is_superuser:
        vcheck = None
        echeck = None
    else:
        vcheck_instance = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=0).first()
        vcheck = vcheck_instance.item_id_id if vcheck_instance else None
        echeck_instance = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=1).first()
        echeck = echeck_instance.item_id_id if echeck_instance else None
        admin = ''
    if request.user.is_superuser:
        admin = Permissions.objects.all()
    perinfo = Perms.objects.filter(user_id=request.user.id)
    if request.method == 'POST':
        oid = request.POST.get('oid')
        pid = request.POST.get('pid')
        oquaid = request.POST.get('oquaid')
        opqid = request.POST.get('opqid')

        if oid is not None and pid is not None and oquaid is not None and opqid is not None:
            oid = int(oid)
            pid = int(pid)
            oquaid = int(oquaid)
            opqid = int(opqid)

            try:
                # Fetch the product and order objects
                product_info = Products.objects.get(id=pid)
                order_info = Orders.objects.get(id=oid)

                if oquaid <= opqid:
                    opqid = opqid - oquaid

                    # Update the product quantity
                    product_info.quantity = opqid
                    product_info.save()

                    # Mark the order as accepted
                    order_info.accept = 1
                    order_info.save()

                    messages.success(request, 'Order has been accepted successfully',extra_tags='alert-success')
                else:
                    messages.error(request, 'Insufficient quantity of goods to confirm the order',extra_tags='alert-danger')
            except (Products.DoesNotExist, Orders.DoesNotExist):
                messages.error(request, 'Product or order with the specified ID was not found',extra_tags='alert-danger')
        else:
            messages.error(request, 'Invalid or missing parameters in the request',extra_tags='alert-danger')
    else:
        messages.error(request, 'Invalid request method',extra_tags='alert-danger')

    data = {'settings':settings,'echeck': echeck, 'vcheck': vcheck, 'perinfo': perinfo, 'admin': admin}
    if 'cancel' in request.POST:
        return HttpResponseRedirect(reverse('orders'), {'data': data})
    if request.user.is_superuser:
        return HttpResponseRedirect(reverse('orders'))  # Redirect to 'brands' for superuser
    else:
        return HttpResponseRedirect(reverse('orders', args=[check, order_id])) 


@login_required
def cancel_order(request,order_id,check=None):
    settings = Settings.objects.first()
    if request.user.is_superuser:
        vcheck = None
        echeck = None
    else:
        vcheck_instance = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=0).first()
        vcheck = vcheck_instance.item_id_id if vcheck_instance else None
        echeck_instance = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=1).first()
        echeck = echeck_instance.item_id_id if echeck_instance else None
        admin = ''
    if request.user.is_superuser:
        admin = Permissions.objects.all()

    perinfo = Perms.objects.filter(user_id=request.user.id)
    if request.method == 'POST':
        oid = request.POST.get('oid')
        pid = request.POST.get('pid')
        oquaid = request.POST.get('oquaid')
        opqid = request.POST.get('opqid')

        if oid is not None and pid is not None and oquaid is not None and opqid is not None:
            oid = int(oid)
            pid = int(pid)
            oquaid = int(oquaid)
            opqid = int(opqid)

            # Отменяем заказ
            info = Products.objects.get(id=pid)
            info.quantity = info.quantity + oquaid
            info.save()

            info = Orders.objects.get(id=oid)
            info.accept = 0
            info.save()

            messages.info(request, 'Order has been canceled successfully',extra_tags='alert-success')
        else:
            messages.error(request, 'Invalid or missing parameters in the request',extra_tags='alert-danger')
    else:
        messages.error(request, 'Invalid request method',extra_tags='alert-danger')

    data = {'settings':settings,'echeck': echeck, 'vcheck': vcheck, 'perinfo': perinfo, 'admin': admin}
    if 'cancel' in request.POST:
        return HttpResponseRedirect(reverse('orders'), {'data': data})
    if request.user.is_superuser:
        return HttpResponseRedirect(reverse('orders'))  # Redirect to 'brands' for superuser
    else:
        return HttpResponseRedirect(reverse('orders', args=[check, order_id])) 



def accept_task(request, planner_id ,check=None):
    settings = Settings.objects.first()
    if request.user.is_superuser:
        vcheck = None
        echeck = None
    else:
        vcheck_instance = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=0).first()
        vcheck = vcheck_instance.item_id_id if vcheck_instance else None
        echeck_instance = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=1).first()
        echeck = echeck_instance.item_id_id if echeck_instance else None
        admin = ''
    if request.user.is_superuser:
        admin = Permissions.objects.all()

    perinfo = Perms.objects.filter(user_id=request.user.id)
    if request.method == 'GET':
        try:
            planner = Planner.objects.get(id=planner_id)
            planner.accept = 1
            planner.save()
            messages.info(request, 'Task has been accepted successfully',extra_tags='alert-success')
        except Planner.DoesNotExist:
            messages.error(request, 'Invalid or missing parameters in the request',extra_tags='alert-danger')
    else:
        messages.error(request, 'Invalid request method.',extra_tags='alert-danger')

    data = {'settings':settings,'echeck': echeck, 'vcheck': vcheck, 'perinfo': perinfo, 'admin': admin}
    if 'cancel' in request.POST:
        return HttpResponseRedirect(reverse('planner'), {'data': data})
    if request.user.is_superuser:
        return HttpResponseRedirect(reverse('planner'))  # Redirect to 'brands' for superuser
    else:
        return HttpResponseRedirect(reverse('planner', args=[check, planner_id])) 

@login_required
def cancel_task(request, planner_id, check=None):
    settings = Settings.objects.first()
    if request.user.is_superuser:
        vcheck = None
        echeck = None
    else:
        vcheck_instance = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=0).first()
        vcheck = vcheck_instance.item_id_id if vcheck_instance else None
        echeck_instance = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=1).first()
        echeck = echeck_instance.item_id_id if echeck_instance else None
        admin = ''
    if request.user.is_superuser:
        admin = Permissions.objects.all()
    perinfo = Perms.objects.filter(user_id=request.user.id)
    if request.method == 'GET':
        try:
            planner = Planner.objects.get(id=planner_id)
            planner.accept = 0
            planner.save()
            messages.info(request, 'Task has been canceled successfully',extra_tags='alert-success')
        except Planner.DoesNotExist:
            messages.error(request, 'Invalid or missing parameters in the request',extra_tags='alert-danger')
    else:
        messages.error(request, 'Invalid request method.',extra_tags='alert-danger')

    data = {'settings':settings,'echeck': echeck, 'vcheck': vcheck, 'perinfo': perinfo, 'admin': admin}
    if 'cancel' in request.POST:
        return HttpResponseRedirect(reverse('planner'), {'data': data})
    if request.user.is_superuser:
        return HttpResponseRedirect(reverse('planner'))  # Redirect to 'brands' for superuser
    else:
        return HttpResponseRedirect(reverse('planner', args=[check, planner_id])) 

#REGISTER VIEW
from django.contrib import messages
from django.shortcuts import render, redirect

def register(request):
    if request.method == 'POST':
        name = request.POST['name']
        surname = request.POST['surname']
        username = request.POST['email']
        phone = request.POST['phone']
        email = request.POST['email']
        password = request.POST['password']
        confirm_password = request.POST['cpassword']

        if password == confirm_password:
            if CustomUser.objects.filter(username=username).exists():
                messages.error(request, 'This username is already in use!',extra_tags='alert-warning')
            else:
                user = CustomUser.objects.create_user(username=username, email=email, password=password, first_name=name, last_name=surname, phone=phone,is_blocked=1)
                messages.success(request, 'Registration successful',extra_tags='alert-success')
                return redirect('login')
        else:
            messages.error(request, 'Passwords do not match.',extra_tags='alert-danger')

    return render(request, 'register.html')








from django.contrib.auth import authenticate, login as auth_login, logout as auth_logout

def login(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        
        if username == '' or password == '':
            messages.info(request, 'Please enter username and password.', extra_tags='alert-warning')
            return redirect('login')

        # Authenticate using the CustomUser model
        user = authenticate(request, username=username, password=password)

        if user:
            if user.is_blocked:
                messages.info(request, 'Your account is currently blocked, please write your message in the contact form.', extra_tags='alert-danger')
            else:
                auth_login(request, user)
                return redirect('profile')
        else:
            messages.info(request, 'Login or password is incorrect.', extra_tags='alert-danger')

    return render(request, 'login.html')

def logout(request):
    auth_logout(request)
    return redirect('login')

#EXCEL  VIEW


@login_required
def export_brands_to_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active

    thin_border = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin'))

    brands_data = Brands.objects.order_by('-id')


    headers = ['№', 'User ID', 'Brand']
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.border = thin_border  
        cell.font = cell.font.copy(bold=True)  

    for row_num, brand in enumerate(brands_data, start=2):
        ws.cell(row=row_num, column=1, value=row_num - 1).border = thin_border
        ws.cell(row=row_num, column=2, value=brand.user_id).border = thin_border
        ws.cell(row=row_num, column=3, value=brand.brand).border = thin_border

    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="brands.xlsx"'

    wb.save(response)

    return response

@login_required
def export_clients_to_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active

    thin_border = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin'))

    clients_data = Clients.objects.order_by('-id')

    headers = ['№', 'User ID', 'Name', 'Surname', 'Email', 'Telephone', 'Company']
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.border = thin_border  
        cell.font = cell.font.copy(bold=True)  

    for row_num, client in enumerate(clients_data, start=2):
        ws.cell(row=row_num, column=1, value=row_num - 1).border = thin_border
        ws.cell(row=row_num, column=2, value=client.user_id).border = thin_border
        ws.cell(row=row_num, column=3, value=client.name).border = thin_border
        ws.cell(row=row_num, column=4, value=client.surname).border = thin_border
        ws.cell(row=row_num, column=5, value=client.email).border = thin_border
        ws.cell(row=row_num, column=6, value=client.telephone).border = thin_border
        ws.cell(row=row_num, column=7, value=client.company).border = thin_border

    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="clients.xlsx"'

    wb.save(response)

    return response

@login_required
def export_departments_to_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active

    thin_border = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin'))

    departments_data = Departments.objects.order_by('-id')

    headers = ['№', 'User ID', 'Department']
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.border = thin_border  
        cell.font = cell.font.copy(bold=True)  

    for row_num, department in enumerate(departments_data, start=2):
        ws.cell(row=row_num, column=1, value=row_num - 1).border = thin_border
        ws.cell(row=row_num, column=2, value=department.user_id).border = thin_border
        ws.cell(row=row_num, column=3, value=department.department).border = thin_border

    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="departments.xlsx"'

    wb.save(response)

    return response

@login_required
def export_expense_to_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active

    thin_border = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin'))

    expenses_data = Expense.objects.order_by('-id')

    headers = ['№', 'User ID', 'Appointment', 'Amount']
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.border = thin_border 
        cell.font = cell.font.copy(bold=True)  

    for row_num, expense in enumerate(expenses_data, start=2):
        ws.cell(row=row_num, column=1, value=row_num - 1).border = thin_border
        ws.cell(row=row_num, column=2, value=expense.user_id).border = thin_border
        ws.cell(row=row_num, column=3, value=expense.appointment).border = thin_border
        ws.cell(row=row_num, column=4, value=expense.amount).border = thin_border

    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="expense.xlsx"'

    wb.save(response)

    return response

@login_required
def export_orders_to_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active

    thin_border = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin'))

    orders_data = Orders.objects.order_by('-id')

    headers = ['№', 'Product', 'Client', 'Purchase', 'Sale', 'Amount', 'Quantity']
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.border = thin_border 
        cell.font = cell.font.copy(bold=True) 

    for row_num, order in enumerate(orders_data, start=2):
        product_info = f"{order.product_id.brand_id.brand}-{order.product_id.name} [{order.product_id.quantity}]"
        client_info = f"{order.client_id.name} {order.client_id.surname}"

        ws.cell(row=row_num, column=1, value=row_num - 1).border = thin_border
        ws.cell(row=row_num, column=2, value=product_info).border = thin_border
        ws.cell(row=row_num, column=3, value=client_info).border = thin_border
        ws.cell(row=row_num, column=4, value=order.product_id.purchase).border = thin_border
        ws.cell(row=row_num, column=5, value=order.product_id.sale).border = thin_border
        ws.cell(row=row_num, column=6, value=order.product_id.quantity).border = thin_border
        ws.cell(row=row_num, column=7, value=order.quantity).border = thin_border

    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="orders.xlsx"'

    wb.save(response)

    return response

@login_required
def export_planner_to_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active

    thin_border = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin'))

    planner_data = Planner.objects.order_by('-id')

    headers = ['№', 'User ID', 'Staff', 'Task', 'Date', 'Time', 'Remaining Time']
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.border = thin_border  # Установите границу для заголовка
        cell.font = cell.font.copy(bold=True)  # Сделайте текст заголовка жирным

    for row_num, planner in enumerate(planner_data, start=2):
        ws.cell(row=row_num, column=1, value=row_num - 1).border = thin_border
        ws.cell(row=row_num, column=2, value=planner.user_id).border = thin_border
        ws.cell(row=row_num, column=3, value=f"{planner.staff_id.name} {planner.staff_id.surname}").border = thin_border
        ws.cell(row=row_num, column=4, value=planner.task).border = thin_border
        ws.cell(row=row_num, column=5, value=planner.dtask.strftime('%Y-%m-%d')).border = thin_border
        ws.cell(row=row_num, column=6, value=planner.ttask.strftime('%H:%M:%S')).border = thin_border
        ws.cell(row=row_num, column=7, value=planner.remaining_time).border = thin_border

    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="planner.xlsx"'

    wb.save(response)

    return response

@login_required
def export_positions_to_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active

    thin_border = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin'))

    positions_data = Positions.objects.order_by('-id')

    headers = ['№', 'User ID', 'Department', 'Position']
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.border = thin_border  
        cell.font = cell.font.copy(bold=True)  

    for row_num, position in enumerate(positions_data, start=2):
        ws.cell(row=row_num, column=1, value=row_num - 1).border = thin_border
        ws.cell(row=row_num, column=2, value=position.user_id).border = thin_border
        ws.cell(row=row_num, column=3, value=position.department_id.department).border = thin_border
        ws.cell(row=row_num, column=4, value=position.positions).border = thin_border

    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="positions.xlsx"'

    wb.save(response)

    return response

@login_required
def export_products_to_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active

    thin_border = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin'))

    products_data = Products.objects.all()

    headers = ['№', 'User ID', 'Firm', 'Brand', 'Product Name', 'Purchase Price', 'Sale Price', 'Quantity', 'Profit']
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.border = thin_border  
        cell.font = cell.font.copy(bold=True)  

    for row_num, product in enumerate(products_data, start=2):
        ws.cell(row=row_num, column=1, value=row_num - 1).border = thin_border
        ws.cell(row=row_num, column=2, value=request.user.id).border = thin_border
        ws.cell(row=row_num, column=3, value=product.supplier_id.firm).border = thin_border
        ws.cell(row=row_num, column=4, value=product.brand_id.brand).border = thin_border
        ws.cell(row=row_num, column=5, value=product.name).border = thin_border
        ws.cell(row=row_num, column=6, value=product.purchase).border = thin_border
        ws.cell(row=row_num, column=7, value=product.sale).border = thin_border
        ws.cell(row=row_num, column=8, value=product.quantity).border = thin_border
        ws.cell(row=row_num, column=9, value=product.qazanc).border = thin_border

    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="products.xlsx"'

    wb.save(response)

    return response

@login_required
def export_staff_to_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active

    thin_border = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin'))

    staff_data = Staff.objects.all()

    headers = ['№', 'User ID', 'Title', 'Name', 'Surname', 'Email', 'Telephone', 'Salary', 'Work Start Date']
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.border = thin_border
        cell.font = cell.font.copy(bold=True)

    for row_num, staff in enumerate(staff_data, start=2):
        position_info = f"{staff.position_id.department_id.department}/{staff.position_id.positions}"
        ws.cell(row=row_num, column=1, value=row_num - 1).border = thin_border
        ws.cell(row=row_num, column=2, value=staff.user_id).border = thin_border
        ws.cell(row=row_num, column=3, value=position_info).border = thin_border
        ws.cell(row=row_num, column=4, value=staff.name).border = thin_border
        ws.cell(row=row_num, column=5, value=staff.surname).border = thin_border
        ws.cell(row=row_num, column=6, value=staff.email).border = thin_border
        ws.cell(row=row_num, column=7, value=staff.telephone).border = thin_border
        ws.cell(row=row_num, column=8, value=staff.salary).border = thin_border
        ws.cell(row=row_num, column=9, value=staff.work.strftime('%Y-%m-%d')).border = thin_border

    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="staff.xlsx"'

    wb.save(response)

    return response

@login_required
def export_suppliers_to_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active

    thin_border = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin'))

    suppliers_data = Suppliers.objects.all()

    headers = ['№', 'User ID', 'Firm', 'Name', 'Surname', 'Email', 'Telephone', 'Address']
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.border = thin_border
        cell.font = cell.font.copy(bold=True)

    for row_num, supplier in enumerate(suppliers_data, start=2):
        ws.cell(row=row_num, column=1, value=row_num - 1).border = thin_border
        ws.cell(row=row_num, column=2, value=supplier.user_id).border = thin_border
        ws.cell(row=row_num, column=3, value=supplier.firm).border = thin_border
        ws.cell(row=row_num, column=4, value=supplier.name).border = thin_border
        ws.cell(row=row_num, column=5, value=supplier.surname).border = thin_border
        ws.cell(row=row_num, column=6, value=supplier.email).border = thin_border
        ws.cell(row=row_num, column=7, value=supplier.telephone).border = thin_border
        ws.cell(row=row_num, column=8, value=supplier.address).border = thin_border

    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="suppliers.xlsx"'

    wb.save(response)

    return response
#END EXCEL

#delete selected view
@login_required
def delete_selected_brands(request,check=None,user_id=None):
    if request.method == 'POST':
        selected_brands = request.POST.getlist('secim[]')
        
        if selected_brands:
            Brands.objects.filter(id__in=selected_brands).delete()
            messages.success(request, 'Selected brands have been deleted successfully!',extra_tags='alert-success')
        else:
            messages.error(request, 'No brands selected for deletion.',extra_tags='alert-warning')

    user = request.user
    user_id=request.user.id
    user_photo=user.photo
    first_name = user.first_name
    last_name = user.last_name
    email = user.email
    info = Brands.objects.order_by('-id')
    say = Brands.objects.count()
        #STAT START
    osay = Orders.objects.count()
    bsay = Brands.objects.count()
    psay = Products.objects.count()
        #STAT END

    settings = Settings.objects.first()
    if request.user.is_superuser:
        vcheck = None
        echeck = None
    else:
        vcheck = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=0).count()
        echeck = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=1).count()
    admin = ''
    if request.user.is_superuser:
        admin = Permissions.objects.all()

    perinfo = Perms.objects.filter(user_id=request.user.id)
    data = {'user_id':user_id,'admin':admin,'perinfo':perinfo,'vcheck':vcheck,'echeck':echeck,'settings':settings,'info':info,'say':say,'first_name':first_name,'last_name':last_name,'email':email,'osay':osay,'bsay':bsay,'psay':psay,'user_photo':user_photo}

    return render(request, 'brands.html',data)

@login_required
def delete_selected_clients(request, check=None, user_id=None):
    if request.method == 'POST':
        selected_clients = request.POST.getlist('secim[]')

        if selected_clients:
            Clients.objects.filter(id__in=selected_clients).delete()
            messages.success(request, 'Selected clients have been deleted successfully!', extra_tags='alert-success')
        else:
            messages.error(request, 'No clients selected for deletion.', extra_tags='alert-warning')

    info = Clients.objects.order_by('-id')
    user = request.user
    user_id = request.user.id
    user_photo = user.photo
    first_name = user.first_name
    last_name = user.last_name
    email = user.email

    say = Clients.objects.count()
        #STAT START
    osay = Orders.objects.count()
    csay = Clients.objects.count()
    psay = Products.objects.count()
    settings = Settings.objects.first()
    
    if request.user.is_superuser:
        vcheck = None
        echeck = None
    else:
        vcheck_instance = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=0).first()
        vcheck = vcheck_instance.item_id_id if vcheck_instance else None
        echeck_instance = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=1).first()
        echeck = echeck_instance.item_id_id if echeck_instance else None
        admin = ''
    if request.user.is_superuser:
        admin = Permissions.objects.all()

    perinfo = Perms.objects.filter(user_id=request.user.id)

    data = {
        'user_id': user_id,
        'admin': admin,
        'perinfo': perinfo,
        'vcheck': vcheck,
        'echeck': echeck,
        'settings': settings,
        'info': info,
        'say': say,
        'first_name': first_name,
        'last_name': last_name,
        'email': email,
        'osay': osay,
        'csay': csay,
        'psay': psay,
        'user_photo': user_photo,
    }

    return render(request, 'clients.html', data)

@login_required
def delete_selected_departments(request,check=None,user_id=None):
    if request.method == 'POST':
        selected_departments = request.POST.getlist('secim[]')
        
        if selected_departments:
            Departments.objects.filter(id__in=selected_departments).delete()
            messages.success(request, 'Selected departments have been deleted successfully!',extra_tags='alert-success')
        else:
            messages.error(request, 'No departments selected for deletion.',extra_tags='alert-warning')

    info = Departments.objects.order_by('-id')
    say = Departments.objects.count()
    user = request.user
    user_id = request.user.id
    first_name = user.first_name
    last_name = user.last_name
    email = user.email
    user_photo = user.photo
    #STAT START
    dsay = Departments.objects.count()
    ssay = Staff.objects.count()
    psay = Positions.objects.count()
    # STAT END
    settings = Settings.objects.first()
    if request.user.is_superuser:
        vcheck = None
        echeck = None
    else:
        vcheck_instance = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=0).first()
        vcheck = vcheck_instance.item_id_id if vcheck_instance else None
        echeck_instance = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=1).first()
        echeck = echeck_instance.item_id_id if echeck_instance else None
        admin = ''
    if request.user.is_superuser:
        admin = Permissions.objects.all()

    perinfo = Perms.objects.filter(user_id=request.user.id)
    
    data = {'user_id':user_id,'admin':admin,'perinfo':perinfo,'vcheck':vcheck,'echeck':echeck,'settings':settings,'info':info,'say':say,'first_name':first_name,'last_name':last_name,'email':email,'dsay':dsay,'ssay':ssay,'psay':psay,'user_photo':user_photo}
    return render(request, 'departments.html',data)

@login_required
def delete_selected_expense(request,check=None,user_id=None):
    if request.method == 'POST':
        selected_expense = request.POST.getlist('secim[]')
        if selected_expense:
            Expense.objects.filter(id__in=selected_expense).delete()
            messages.success(request, 'Selected expenses have been deleted successfully!',extra_tags='alert-success')
        else:
            messages.error(request, 'No valid expenses selected for deletion.',extra_tags='alert-warning')

    info = Expense.objects.order_by('-id')
    say = Expense.objects.count()
    user = request.user
    user_id = request.user.id
    first_name = user.first_name
    last_name = user.last_name
    email = user.email
    user_photo = user.photo
    #STAT START
    esay = Expense.objects.count()
    # STAT END
    settings = Settings.objects.first()
    if request.user.is_superuser:
        vcheck = None
        echeck = None
    else:
        vcheck_instance = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=0).first()
        vcheck = vcheck_instance.item_id_id if vcheck_instance else None
        echeck_instance = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=1).first()
        echeck = echeck_instance.item_id_id if echeck_instance else None
        admin = ''
    if request.user.is_superuser:
        admin = Permissions.objects.all()

    perinfo = Perms.objects.filter(user_id=request.user.id)

    data = {'user_id':user_id,'admin':admin,'perinfo':perinfo,'vcheck':vcheck,'echeck':echeck,'settings':settings,'info':info,'say':say,'first_name':first_name,'last_name':last_name,'email':email,'esay':esay,'user_photo':user_photo}
    return render(request, 'expense.html',data)

@login_required
def delete_selected_orders(request,check=None, user_id=None):
    if request.method == 'POST':
        selected_order = request.POST.getlist('secim[]')
        if selected_order:
            Orders.objects.filter(id__in=selected_order).delete()
            messages.success(request, 'Selected orders have been deleted successfully!',extra_tags='alert-success')
        else:
            messages.error(request, 'No valid orders selected for deletion.',extra_tags='alert-warning')

    info = Orders.objects.order_by('-id')
    say = Orders.objects.count()
    user = request.user
    user_id = request.user.id
    first_name = user.first_name
    last_name = user.last_name
    email = user.email
    active_count = Orders.objects.filter(accept=0).count()
    completed_count = Orders.objects.filter(accept=1).count()
    user_photo = user.photo
    #STAT START
    osay = Orders.objects.count()
    bsay = Brands.objects.count()
    csay = Clients.objects.count()
    psay = Products.objects.count()
    # STAT END
    cinfo = Clients.objects.all()
    pinfo = Products.objects.all()
    settings = Settings.objects.first()
    if request.user.is_superuser:
        vcheck = None
        echeck = None
    else:
        vcheck_instance = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=0).first()
        vcheck = vcheck_instance.item_id_id if vcheck_instance else None
        echeck_instance = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=1).first()
        echeck = echeck_instance.item_id_id if echeck_instance else None
        admin = ''
    if request.user.is_superuser:
        admin = Permissions.objects.all()

    perinfo = Perms.objects.filter(user_id=request.user.id)

    data = {'active_count':active_count,'completed_count':completed_count,'user_id':user_id,'admin':admin,'perinfo':perinfo,'vcheck':vcheck,'echeck':echeck,'settings':settings,'info':info,'say':say,'first_name':first_name,'last_name':last_name,'email':email,'osay':osay,'bsay':bsay,'csay':csay,'psay':psay,'cinfo':cinfo,'pinfo':pinfo,'user_photo':user_photo}
    return render(request, 'orders.html', data)

@login_required
def delete_selected_planner(request,check=None,user_id=None):
    if request.method == 'POST':
        selected_planner = request.POST.getlist('secim[]')
        if selected_planner:
            Planner.objects.filter(id__in=selected_planner).delete()
            messages.success(request, 'Selected planner have been deleted successfully!',extra_tags='alert-success')
        else:
            messages.error(request, 'No valid planner selected for deletion.',extra_tags='alert-warning')

    info = Planner.objects.order_by('-id')
    say = Planner.objects.count()
    active_count = Planner.objects.filter(accept=0).count()
    completed_count = Planner.objects.filter(accept=1).count()
    user = request.user
    user_id = request.user.id
    first_name = user.first_name
    last_name = user.last_name
    email = user.email
    user_photo = user.photo
    #STAT START
    osay = Orders.objects.count()
    bsay = Brands.objects.count()
    csay = Clients.objects.count()
    psay = Products.objects.count()
    sinfo = Staff.objects.all()
    settings = Settings.objects.first()
    if request.user.is_superuser:
        vcheck = None
        echeck = None
    else:
        vcheck_instance = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=0).first()
        vcheck = vcheck_instance.item_id_id if vcheck_instance else None
        echeck_instance = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=1).first()
        echeck = echeck_instance.item_id_id if echeck_instance else None
        admin = ''
    if request.user.is_superuser:
        admin = Permissions.objects.all()

    perinfo = Perms.objects.filter(user_id=request.user.id)

    data = {'active_count':active_count,'completed_count':completed_count,'user_id':user_id,'admin':admin,'perinfo':perinfo,'vcheck':vcheck,'echeck':echeck,'settings':settings,'info':info,'say':say,'first_name':first_name,'last_name':last_name,'email':email,'osay':osay,'bsay':bsay,'csay':csay,'psay':psay,'sinfo':sinfo,'user_photo':user_photo}
    return render(request, 'planner.html',data)

@login_required
def delete_selected_positions(request,check=None,user_id=None):
    if request.method == 'POST':
        selected_positions = request.POST.getlist('secim[]')
        if selected_positions:
            Positions.objects.filter(id__in=selected_positions).delete()
            messages.success(request, 'Selected positions have been deleted successfully!',extra_tags='alert-success')
        else:
            messages.error(request, 'No valid positions selected for deletion.',extra_tags='alert-warning')

    info = Positions.objects.order_by('-id')
    say = Positions.objects.count()
    user = request.user
    first_name = user.first_name
    last_name = user.last_name
    email = user.email
    user_photo = user.photo
    #STAT START
    dsay = Departments.objects.count()
    psay = Positions.objects.count()
    ssay = Staff.objects.count()
    dinfo = Departments.objects.all()
    settings = Settings.objects.first()
    if request.user.is_superuser:
        vcheck = None
        echeck = None
    else:
        vcheck_instance = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=0).first()
        vcheck = vcheck_instance.item_id_id if vcheck_instance else None
        echeck_instance = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=1).first()
        echeck = echeck_instance.item_id_id if echeck_instance else None
        admin = ''
    if request.user.is_superuser:
        admin = Permissions.objects.all()

    perinfo = Perms.objects.filter(user_id=request.user.id)

    data = {'user_id':user_id,'admin':admin,'perinfo':perinfo,'vcheck':vcheck,'echeck':echeck,'settings':settings,'info':info,'say':say,'first_name':first_name,'last_name':last_name,'email':email,'dsay':dsay,'ssay':ssay,'psay':psay,'dinfo':dinfo,'user_photo':user_photo}
    return render(request, 'positions.html',data)

@login_required
def delete_selected_products(request,check=None,user_id=None):
    if request.method == 'POST':
        selected_products = request.POST.getlist('secim[]')
        if selected_products:
            Products.objects.filter(id__in=selected_products).delete()
            messages.success(request, 'Selected products have been deleted successfully!',extra_tags='alert-success')
        else:
            messages.error(request, 'No valid products selected for deletion.',extra_tags='alert-warning')

    info = Products.objects.order_by('-id')
    say = Products.objects.count()
    user = request.user
    user_id = request.user.id
    first_name = user.first_name
    last_name = user.last_name
    email = user.email
    user_photo = user.photo
    #STAT START
    osay = Orders.objects.count()
    bsay = Brands.objects.count()
    csay = Clients.objects.count()
    psay = Products.objects.count()
    sinfo = Suppliers.objects.all()
    binfo = Brands.objects.all()
    settings = Settings.objects.first()
    if request.user.is_superuser:
        vcheck = None
        echeck = None
    else:
        vcheck_instance = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=0).first()
        vcheck = vcheck_instance.item_id_id if vcheck_instance else None
        echeck_instance = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=1).first()
        echeck = echeck_instance.item_id_id if echeck_instance else None
        admin = ''
    if request.user.is_superuser:
        admin = Permissions.objects.all()

    perinfo = Perms.objects.filter(user_id=request.user.id)

    data = {'user_id':user_id,'admin':admin,'perinfo':perinfo,'vcheck':vcheck,'echeck':echeck,'settings':settings,'info':info,'say':say,'first_name':first_name,'last_name':last_name,'email':email,'osay':osay,'bsay':bsay,'csay':csay,'psay':psay,'sinfo':sinfo,'binfo':binfo,'user_photo':user_photo}
    return render(request, 'products.html',data)

@login_required
def delete_selected_staff(request,check=None,user_id=None):
    if request.method == 'POST':
        selected_staff = request.POST.getlist('secim[]')
        if selected_staff:
            Staff.objects.filter(id__in=selected_staff).delete()
            messages.success(request, 'Selected staff have been deleted successfully!',extra_tags='alert-success')
        else:
            messages.error(request, 'No valid staff selected for deletion.',extra_tags='alert-warning')

    info = Staff.objects.order_by('-id')
    say = Staff.objects.count()
    user = request.user
    user_id = request.user.id
    first_name = user.first_name
    last_name = user.last_name
    email = user.email
    user_photo = user.photo
    #STAT START
    dsay = Departments.objects.count()
    ssay = Staff.objects.count()
    psay = Products.objects.count()

    dinfo = Departments.objects.all()
    pinfo = Positions.objects.all()
    settings = Settings.objects.first()
    if request.user.is_superuser:
        vcheck = None
        echeck = None
    else:
        vcheck_instance = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=0).first()
        vcheck = vcheck_instance.item_id_id if vcheck_instance else None
        echeck_instance = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=1).first()
        echeck = echeck_instance.item_id_id if echeck_instance else None
        admin = ''
    if request.user.is_superuser:
        admin = Permissions.objects.all()

    perinfo = Perms.objects.filter(user_id=request.user.id)

    data = {'user_id':user_id,'admin':admin,'perinfo':perinfo,'vcheck':vcheck,'echeck':echeck,'settings':settings,'info':info,'say':say,'first_name':first_name,'last_name':last_name,'email':email,'dsay':dsay,'ssay':ssay,'psay':psay,'dinfo':dinfo,'pinfo':pinfo,'user_photo':user_photo}
    return render(request, 'staff.html',data)

@login_required
def delete_selected_suppliers(request,check=None,user_id=None):
    if request.method == 'POST':
        selected_suppliers = request.POST.getlist('secim[]')
        if selected_suppliers:
            Suppliers.objects.filter(id__in=selected_suppliers).delete()
            messages.success(request, 'Selected suppliers have been deleted successfully!',extra_tags='alert-success')
        else:
            messages.error(request, 'No valid suppliers selected for deletion.',extra_tags='alert-warning')

    info = Suppliers.objects.order_by('-id')
    say = Suppliers.objects.count()
    user = request.user
    user_id = request.user.id
    first_name = user.first_name
    last_name = user.last_name
    email = user.email
    user_photo = user.photo
    #STAT START
    osay = Orders.objects.count()
    bsay = Brands.objects.count()
    csay = Clients.objects.count()
    psay = Products.objects.count()
    settings = Settings.objects.first()
    if request.user.is_superuser:
        vcheck = None
        echeck = None
    else:
        vcheck_instance = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=0).first()
        vcheck = vcheck_instance.item_id_id if vcheck_instance else None
        echeck_instance = Perms.objects.filter(item_id_id=check, user_id=request.user.id, item_value=1).first()
        echeck = echeck_instance.item_id_id if echeck_instance else None
        admin = ''
    if request.user.is_superuser:
        admin = Permissions.objects.all()

    perinfo = Perms.objects.filter(user_id=request.user.id)

    data = {'user_id':user_id,'admin':admin,'perinfo':perinfo,'vcheck':vcheck,'echeck':echeck,'settings':settings,'info':info,'say':say,'first_name':first_name,'last_name':last_name,'email':email,'osay':osay,'bsay':bsay,'csay':csay,'psay':psay,'user_photo':user_photo}
    return render(request, 'suppliers.html',data)


#end selected


